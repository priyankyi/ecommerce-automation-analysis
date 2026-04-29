from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.marketplaces.flipkart.flipkart_utils import (
    ANALYSIS_JSON_PATH,
    FSN_BRIDGE_PATH,
    NORMALIZATION_LOG_PATH,
    NORMALIZATION_STATE_PATH,
    NORMALIZED_ADS_PATH,
    NORMALIZED_LISTINGS_PATH,
    NORMALIZED_ORDERS_PATH,
    NORMALIZED_PNL_PATH,
    NORMALIZED_RETURNS_PATH,
    NORMALIZED_SALES_TAX_PATH,
    NORMALIZED_SETTLEMENTS_PATH,
    RAW_INPUT_DIR,
    TARGET_FSN_PATH,
    ensure_directories,
    format_decimal,
    clean_fsn,
    get_cell,
    highest_priority_fsn,
    load_json,
    normalize_text,
    now_iso,
    parse_float,
    read_workbook_rows,
    load_synonyms,
    load_report_patterns,
    detect_header_row,
    list_input_files,
    select_best_sheet_across_files,
    write_csv,
    append_csv_log,
    dedupe_dict_rows,
    save_run_state,
    build_status_payload,
)

LOG_HEADERS = [
    "timestamp",
    "output_file",
    "source_file",
    "sheet_name",
    "rows_read",
    "rows_written",
    "rows_before_filter",
    "rows_after_fsn_filter",
    "unmapped_rows",
    "reason_if_zero",
    "status",
    "message",
]
LISTING_HEADERS = ["FSN", "Seller_SKU", "Product_Title", "Category", "Listing_Status", "Inactive_Reason", "MRP", "Selling_Price", "Stock", "Bank_Settlement", "Listing_Quality", "Package_Length", "Package_Breadth", "Package_Height", "Package_Weight", "HSN", "Tax_Code", "Source_File", "Mapping_Confidence"]
ORDER_HEADERS = ["FSN", "Order_ID", "Order_Item_ID", "Seller_SKU", "Product_Title", "Order_Date", "Quantity", "Selling_Price", "Order_Status", "Dispatch_Date", "Delivery_Date", "Cancellation_Status", "Source_File", "Mapping_Confidence"]
RETURN_HEADERS = ["FSN", "Order_ID", "Order_Item_ID", "Seller_SKU", "Return_ID", "Return_Date", "Return_Type", "Return_Reason", "Return_Status", "Reverse_Shipment_Status", "Source_File", "Mapping_Confidence"]
SETTLEMENT_HEADERS = [
    "FSN",
    "Order_ID",
    "Order_Item_ID",
    "Settlement_ID",
    "Settlement_Date",
    "Seller_SKU",
    "Gross_Amount",
    "Commission",
    "Fixed_Fee",
    "Collection_Fee",
    "Shipping_Fee",
    "Reverse_Shipping_Fee",
    "GST_On_Fees",
    "TCS",
    "TDS",
    "Refund",
    "Protection_Fund",
    "Adjustments",
    "Net_Settlement",
    "Source_File",
    "Source_Sheet",
    "Mapping_Confidence",
    "Mapping_Issue",
]
PNL_HEADERS = [
    "FSN",
    "Order_ID",
    "Order_Item_ID",
    "Seller_SKU",
    "Flipkart_Net_Earnings",
    "Flipkart_Margin",
    "Flipkart_Expenses",
    "Amount_Settled",
    "Amount_Pending",
    "Source_File",
    "Source_Sheet",
    "Mapping_Confidence",
    "Mapping_Issue",
]
SALES_TAX_HEADERS = ["FSN", "Order_ID", "Order_Item_ID", "Invoice_ID", "Invoice_Date", "HSN", "Taxable_Value", "IGST", "CGST", "SGST", "TCS", "TDS", "Event_Type", "Source_File", "Mapping_Confidence"]
ADS_HEADERS = ["FSN", "Campaign_ID", "Campaign_Name", "AdGroup_ID", "AdGroup_Name", "Seller_SKU", "Product_Name", "Views", "Clicks", "Direct_Units_Sold", "Indirect_Units_Sold", "Total_Revenue", "ROI", "Estimated_Ad_Spend", "ROAS", "ACOS", "Source_File", "Mapping_Confidence"]
ADS_HEADERS = ADS_HEADERS[:-1] + ["Mapping_Issue", ADS_HEADERS[-1]]

SINGLE_ROW_LIMIT = 10
MAX_HEADER_SCAN_ROWS = 50
REPORT_TYPES = ["listing", "orders", "returns", "sales_tax", "ads"]
SPECIAL_REPORT_TYPES = ["settlements", "pnl"]
SETTLEMENT_WORKBOOK_NAME = "Settled Transactions.xlsx"
PNL_WORKBOOK_NAME = "PNL.xlsx"


def load_analysis() -> Dict[str, Any]:
    if not ANALYSIS_JSON_PATH.exists():
        raise FileNotFoundError(f"Missing required file: {ANALYSIS_JSON_PATH}")
    return load_json(ANALYSIS_JSON_PATH)


def load_target_fsns() -> Dict[str, Dict[str, Any]]:
    if not TARGET_FSN_PATH.exists():
        raise FileNotFoundError(f"Missing required file: {TARGET_FSN_PATH}")
    rows = read_csv_dicts(TARGET_FSN_PATH)
    return {clean_fsn(row["FSN"]): row for row in rows if clean_fsn(row.get("FSN"))}


def load_bridge_sku_lookup() -> Dict[str, List[str]]:
    if not FSN_BRIDGE_PATH.exists():
        raise FileNotFoundError(f"Missing required file: {FSN_BRIDGE_PATH}")
    rows = read_csv_dicts(FSN_BRIDGE_PATH)
    lookup: Dict[str, List[str]] = defaultdict(list)
    for row in rows:
        sku = normalize_text(row.get("Seller_SKU", ""))
        fsn = clean_fsn(row.get("FSN", ""))
        if sku and fsn:
            lookup[sku].append(fsn)
    return {sku: sorted(set(fsns)) for sku, fsns in lookup.items()}


def read_csv_dicts(path: Path) -> List[Dict[str, str]]:
    import csv

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def get_rows_for_sheet(file_path: Path, sheet_name: str) -> List[List[Any]]:
    workbook_rows = read_workbook_rows(file_path)
    return workbook_rows.get(sheet_name, [])


def value(row: Sequence[Any], columns: Dict[str, Dict[str, Any]], key: str) -> str:
    column = columns.get(key)
    if not column:
        return ""
    return normalize_text(get_cell(row, int(column["index"])))


def sheet_has_content(rows: Sequence[Sequence[Any]]) -> bool:
    for row in rows[:MAX_HEADER_SCAN_ROWS]:
        if any(normalize_text(cell) for cell in row):
            return True
    return False


def detect_sheet_columns(rows: Sequence[Sequence[Any]], synonyms: Dict[str, List[str]]) -> Tuple[int, Dict[str, Dict[str, Any]], List[Dict[str, Any]]]:
    return detect_header_row(rows, synonyms, max_scan_rows=MAX_HEADER_SCAN_ROWS)


def count_clean_column_values(rows: Sequence[Sequence[Any]], columns: Dict[str, Dict[str, Any]], key: str) -> int:
    column = columns.get(key)
    if not column:
        return 0
    index = int(column["index"])
    total = 0
    for row in rows:
        if index < len(row) and clean_fsn(row[index]):
            total += 1
    return total


def collect_clean_column_values(rows: Sequence[Sequence[Any]], columns: Dict[str, Dict[str, Any]], key: str, limit: int = SINGLE_ROW_LIMIT) -> List[str]:
    column = columns.get(key)
    if not column:
        return []
    index = int(column["index"])
    values: List[str] = []
    for row in rows:
        if index < len(row):
            value = clean_fsn(row[index])
            if value:
                values.append(value)
        if len(values) >= limit:
            break
    return values


def column_keys_present(columns: Dict[str, Dict[str, Any]], keys: Sequence[str]) -> bool:
    return any(key in columns for key in keys)


def extract_join_key_notes(columns: Dict[str, Dict[str, Any]]) -> bool:
    return column_keys_present(columns, ["fsn", "order_item_id", "order_id", "sku_id"])


def log_deep_scan_candidate(
    report_type: str,
    file_name: str,
    sheet_name: str,
    header_row_index: int,
    columns: Dict[str, Dict[str, Any]],
    data_rows: Sequence[Sequence[Any]],
) -> None:
    print(
        json.dumps(
            {
                "report_type": report_type,
                "file": file_name,
                "sheet": sheet_name,
                "header_row": header_row_index,
                "rows": len(data_rows),
                "columns": sorted(columns.keys()),
            },
            ensure_ascii=False,
        )
    )


def resolve_fsn_from_row(
    row: Sequence[Any],
    columns: Dict[str, Dict[str, Any]],
    target_fsns: Dict[str, Dict[str, Any]],
    priority_map: Dict[str, Dict[str, Any]],
    order_item_lookup: Dict[str, str],
    order_id_lookup: Dict[str, str],
    bridge_sku_lookup: Dict[str, List[str]],
    allow_ambiguous_sku_choice: bool = True,
) -> Tuple[List[str], str, str]:
    fsn = clean_fsn(value(row, columns, "fsn"))
    sku = value(row, columns, "sku_id")
    order_item_id = clean_fsn(value(row, columns, "order_item_id"))
    order_id = clean_fsn(value(row, columns, "order_id"))

    if fsn and fsn in target_fsns:
        return [fsn], "HIGH", ""

    if order_item_id:
        candidate_fsn = order_item_lookup.get(order_item_id, "")
        if candidate_fsn:
            candidate_fsn = clean_fsn(candidate_fsn)
            return [candidate_fsn], "HIGH", "Mapped through order_item_id"
    if order_id:
        candidate_fsn = order_id_lookup.get(order_id, "")
        if candidate_fsn:
            candidate_fsn = clean_fsn(candidate_fsn)
            return [candidate_fsn], "MEDIUM", "Mapped through order_id"
    if sku:
        candidates = bridge_sku_lookup.get(sku, [])
        if not candidates:
            for target_fsn, target_row in target_fsns.items():
                if normalize_text(target_row.get("SKU_ID")) == sku:
                    candidates.append(target_fsn)
        if len(candidates) == 1:
            return [clean_fsn(candidates[0])], "MEDIUM", "SKU fallback"
        if len(candidates) > 1:
            if not allow_ambiguous_sku_choice:
                return [], "LOW", "SKU maps to multiple FSNs"
            chosen = highest_priority_fsn(candidates, priority_map)
            return [clean_fsn(chosen)] if chosen else [], "LOW", "SKU maps to multiple FSNs"
    return [], "LOW", "Unmapped row"


def build_order_lookup(rows: List[Dict[str, Any]]) -> Tuple[Dict[str, str], Dict[str, str]]:
    order_item_to_fsn: Dict[str, str] = {}
    order_to_fsn: Dict[str, str] = {}
    for row in rows:
        fsn = row.get("FSN", "")
        if row.get("Order_Item_ID"):
            order_item_to_fsn.setdefault(row["Order_Item_ID"], fsn)
        if row.get("Order_ID"):
            order_to_fsn.setdefault(row["Order_ID"], fsn)
    return order_item_to_fsn, order_to_fsn


def as_output_row(base: Dict[str, Any], headers: Sequence[str]) -> Dict[str, Any]:
    return {header: base.get(header, "") for header in headers}


def build_reason(report_type: str, raw_rows: int, rows_after_filter: int, unmapped_rows: int, columns: Dict[str, Dict[str, Any]], join_key_detected: bool) -> str:
    if raw_rows == 0:
        return "no data rows after the header"
    if rows_after_filter > 0:
        return ""
    if report_type == "orders":
        return "Order report FSNs do not overlap target FSN file."
    if report_type == "sales_tax" and not join_key_detected:
        return "Not usable for FSN-level analysis from current file structure"
    if report_type == "returns" and not join_key_detected:
        return "Return report has no usable join key"
    if unmapped_rows > 0:
        return "Rows were detected, but no FSN mapping succeeded"
    if not columns:
        return "Header detection failed"
    return "FSNs were found, but none matched the target FSN list"


def report_file_names(analysis: Dict[str, Any], report_type: str) -> List[str]:
    return sorted({file_info.get("file_name", "") for file_info in analysis.get("files", []) if file_info.get("report_type") == report_type and file_info.get("file_name")})


def print_selection(report_type: str, selection: Dict[str, Any]) -> None:
    selected = selection.get("selected_sheet")
    rejected = selection.get("rejected_sheets", [])
    payload: Dict[str, Any] = {"report_type": report_type}
    if selected:
        payload.update(
            {
                "selected_file": selected.get("file_name", ""),
                "selected_sheet": selected.get("sheet_name", ""),
                "selected_score": selected.get("selection_score", 0.0),
                "header_row_index": selected.get("header_row_index", 0),
                "columns": sorted(selected.get("detected_columns", {}).keys()),
                "raw_row_count": selected.get("raw_row_count", 0),
                "normalized_row_count": selected.get("non_empty_row_count", 0),
                "rejected_sheets": [
                    {
                        "sheet_name": item.get("sheet_name", ""),
                        "reason": item.get("exclusion_reason", ""),
                        "score": item.get("selection_score", 0.0),
                    }
                    for item in rejected
                ],
            }
        )
    else:
        payload.update(
            {
                "selected_file": "",
                "selected_sheet": "",
                "selected_score": 0.0,
                "header_row_index": "",
                "columns": [],
                "raw_row_count": 0,
                "normalized_row_count": 0,
                "rejected_sheets": [
                    {
                        "sheet_name": item.get("sheet_name", ""),
                        "reason": item.get("exclusion_reason", ""),
                        "score": item.get("selection_score", 0.0),
                    }
                    for item in rejected
                ],
            }
        )
    print(json.dumps(payload, ensure_ascii=False))


def read_full_workbook_rows(file_path: Path) -> Dict[str, List[List[Any]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=False, data_only=True)
    sheet_rows: Dict[str, List[List[Any]]] = {}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        sheet_rows[sheet_name] = [list(row) for row in worksheet.iter_rows(values_only=True)]
    return sheet_rows


def flatten_header_rows(rows: Sequence[Sequence[Any]], header_row_indexes: Sequence[int]) -> List[str]:
    max_width = 0
    for row_index in header_row_indexes:
        if row_index < len(rows):
            max_width = max(max_width, len(rows[row_index]))

    headers: List[str] = []
    for col_index in range(max_width):
        parts: List[str] = []
        for row_index in header_row_indexes:
            if row_index >= len(rows):
                continue
            row = rows[row_index]
            if col_index >= len(row):
                continue
            part = normalize_text(row[col_index])
            if part and part not in parts:
                parts.append(part)
        headers.append(" ".join(parts).strip())
    return headers


def build_columns_from_header_rows(rows: Sequence[Sequence[Any]], header_row_indexes: Sequence[int], synonyms: Dict[str, List[str]]) -> Dict[str, Dict[str, Any]]:
    flattened = flatten_header_rows(rows, header_row_indexes)
    _, columns, _ = detect_header_row([flattened], synonyms, max_scan_rows=1)
    return columns


def order_identifier_variants(value: Any) -> List[str]:
    text = normalize_text(value).upper()
    if not text:
        return []
    variants = [text]
    for prefix in ("OI:", "OI", "OD:", "OD"):
        if text.startswith(prefix):
            stripped = text[len(prefix) :].lstrip(":-_ ")
            if stripped and stripped not in variants:
                variants.append(stripped)
    return variants


def lookup_fsn_by_identifier(lookup: Dict[str, str], value: Any) -> str:
    for variant in order_identifier_variants(value):
        fsn = lookup.get(variant, "")
        if fsn:
            return fsn
    return ""


def build_order_lookup(rows: List[Dict[str, Any]]) -> Tuple[Dict[str, str], Dict[str, str]]:
    order_item_to_fsn: Dict[str, str] = {}
    order_to_fsn: Dict[str, str] = {}
    for row in rows:
        fsn = row.get("FSN", "")
        for variant in order_identifier_variants(row.get("Order_Item_ID", "")):
            order_item_to_fsn.setdefault(variant, fsn)
        for variant in order_identifier_variants(row.get("Order_ID", "")):
            order_to_fsn.setdefault(variant, fsn)
    return order_item_to_fsn, order_to_fsn


def resolve_join_fsn(
    row: Sequence[Any],
    columns: Dict[str, Dict[str, Any]],
    target_fsns: Dict[str, Dict[str, Any]],
    priority_map: Dict[str, Dict[str, Any]],
    order_item_lookup: Dict[str, str],
    order_id_lookup: Dict[str, str],
    bridge_sku_lookup: Dict[str, List[str]],
    allow_ambiguous_sku_choice: bool = True,
    allow_sku_fallback: bool = True,
) -> Tuple[List[str], str, str]:
    fsn = clean_fsn(value(row, columns, "fsn"))
    sku = value(row, columns, "sku_id")
    order_item_id = normalize_text(select_sheet_value(row, columns, ["order_item_id", "order_item_id_join"]))
    order_id = normalize_text(value(row, columns, "order_id"))

    if fsn and fsn in target_fsns:
        return [fsn], "HIGH", ""

    if order_item_id:
        candidate_fsn = lookup_fsn_by_identifier(order_item_lookup, order_item_id)
        if candidate_fsn:
            return [clean_fsn(candidate_fsn)], "HIGH", "Mapped through order_item_id"

    if order_id:
        candidate_fsn = lookup_fsn_by_identifier(order_id_lookup, order_id)
        if candidate_fsn:
            return [clean_fsn(candidate_fsn)], "MEDIUM", "Mapped through order_id"

    if allow_sku_fallback and sku:
        candidates = bridge_sku_lookup.get(sku, [])
        if not candidates:
            for target_fsn, target_row in target_fsns.items():
                if normalize_text(target_row.get("SKU_ID")) == sku:
                    candidates.append(target_fsn)
        if len(candidates) == 1:
            return [clean_fsn(candidates[0])], "MEDIUM", "SKU fallback"
        if len(candidates) > 1:
            if not allow_ambiguous_sku_choice:
                return [], "LOW", "SKU maps to multiple FSNs"
            chosen = highest_priority_fsn(candidates, priority_map)
            return [clean_fsn(chosen)] if chosen else [], "LOW", "SKU maps to multiple FSNs"

    return [], "LOW", "Unmapped row"


def select_sheet_value(
    row: Sequence[Any],
    columns: Dict[str, Dict[str, Any]],
    key_options: Sequence[str],
) -> str:
    for key in key_options:
        column = columns.get(key)
        if column and int(column["index"]) < len(row):
            value_text = normalize_text(row[int(column["index"])])
            if value_text:
                return value_text
    return ""


def first_non_blank(*values: Any) -> str:
    for value in values:
        text = normalize_text(value)
        if text:
            return text
    return ""


def map_gst_fee_name(fee_name: str) -> str:
    fee = normalize_text(fee_name).lower()
    if not fee:
        return "Adjustments"
    if "commission" in fee:
        return "Commission"
    if "fixed" in fee:
        return "Fixed_Fee"
    if "collection" in fee:
        return "Collection_Fee"
    if "reverse shipping" in fee or "rto shipping" in fee:
        return "Reverse_Shipping_Fee"
    if "shipping" in fee or "forward shipping" in fee:
        return "Shipping_Fee"
    if "refund" in fee:
        return "Refund"
    if "protection" in fee:
        return "Protection_Fund"
    if fee.startswith("tcs") or "tax collected at source" in fee:
        return "TCS"
    if fee.startswith("tds") or "tax deducted at source" in fee:
        return "TDS"
    if "gst" in fee or "tax on fees" in fee:
        return "GST_On_Fees"
    return "Adjustments"


def parse_settlement_workbook(
    file_path: Path,
    target_fsns: Dict[str, Dict[str, Any]],
    order_item_lookup: Dict[str, str],
    order_id_lookup: Dict[str, str],
    bridge_sku_lookup: Dict[str, List[str]],
) -> Tuple[List[Dict[str, Any]], Dict[str, Any], List[Dict[str, Any]], List[Dict[str, Any]]]:
    workbook_rows = read_full_workbook_rows(file_path)
    rows_out: List[Dict[str, Any]] = []
    selected_headers: List[Dict[str, Any]] = []
    sheet_stats: List[Dict[str, Any]] = []
    total_raw_rows = 0
    total_mapped_rows = 0
    total_unmapped_rows = 0
    priority_map = {fsn: row for fsn, row in target_fsns.items()}
    relevant_sheets = [
        "Orders",
        "GST_Details",
        "Non_Order_SPF",
        "Storage_Recall",
        "MP Fee Rebate",
        "TCS_Recovery",
        "TDS",
        "Value Added Services",
        "Google Ads Services",
        "Ads",
    ]

    for sheet_name in relevant_sheets:
        rows = workbook_rows.get(sheet_name, [])
        if not rows:
            sheet_stats.append({"sheet_name": sheet_name, "raw_rows": 0, "mapped_rows": 0, "unmapped_rows": 0, "status": "EMPTY"})
            continue

        if sheet_name == "Orders":
            header_rows = [0, 1]
            data_start = 3
        elif sheet_name in {"GST_Details", "Non_Order_SPF", "Storage_Recall", "MP Fee Rebate", "TCS_Recovery", "TDS", "Ads"}:
            header_rows = [0, 1]
            data_start = 2
        else:
            header_rows = [0, 1]
            data_start = 2

        columns = build_columns_from_header_rows(rows, header_rows, load_synonyms())
        selected_headers.append(
            {
                "file_name": file_path.name,
                "sheet_name": sheet_name,
                "header_row_indexes": header_rows,
                "columns": sorted(columns.keys()),
                "header_preview": flatten_header_rows(rows, header_rows)[:15],
            }
        )

        data_rows = [row for row in rows[data_start:] if any(normalize_text(cell) for cell in row)]
        raw_rows = len(data_rows)
        mapped_rows = 0
        unmapped_rows = 0

        for row in data_rows:
            if sheet_name == "Orders":
                candidate_fsns, confidence, issue = resolve_join_fsn(
                    row,
                    columns,
                    target_fsns,
                    priority_map,
                    order_item_lookup,
                    order_id_lookup,
                    bridge_sku_lookup,
                    allow_ambiguous_sku_choice=True,
                    allow_sku_fallback=True,
                )
                if not candidate_fsns:
                    unmapped_rows += 1
                    continue
                fsn = candidate_fsns[0]
                base_row = {
                    "FSN": fsn,
                    "Order_ID": select_sheet_value(row, columns, ["order_id"]),
                    "Order_Item_ID": select_sheet_value(row, columns, ["order_item_id_join", "order_item_id"]),
                    "Settlement_ID": select_sheet_value(row, columns, ["neft_id"]),
                    "Settlement_Date": select_sheet_value(row, columns, ["settlement_date", "order_date"]),
                    "Seller_SKU": select_sheet_value(row, columns, ["sku_id"]),
                    "Gross_Amount": select_sheet_value(row, columns, ["sale_amount", "gross_amount"]),
                    "Commission": select_sheet_value(row, columns, ["commission"]),
                    "Fixed_Fee": select_sheet_value(row, columns, ["fixed_fee"]),
                    "Collection_Fee": select_sheet_value(row, columns, ["collection_fee"]),
                    "Shipping_Fee": select_sheet_value(row, columns, ["shipping_fee"]),
                    "Reverse_Shipping_Fee": select_sheet_value(row, columns, ["reverse_shipping_fee"]),
                    "GST_On_Fees": select_sheet_value(row, columns, ["gst_on_fees"]),
                    "TCS": select_sheet_value(row, columns, ["tcs"]),
                    "TDS": select_sheet_value(row, columns, ["tds"]),
                    "Refund": select_sheet_value(row, columns, ["refund"]),
                    "Protection_Fund": select_sheet_value(row, columns, ["protection_fund"]),
                    "Adjustments": first_non_blank(
                        select_sheet_value(row, columns, ["adjustments"]),
                        select_sheet_value(row, columns, ["collection_fee"]),
                    ),
                    "Net_Settlement": select_sheet_value(row, columns, ["net_settlement"]),
                    "Source_File": file_path.name,
                    "Source_Sheet": sheet_name,
                    "Mapping_Confidence": confidence,
                    "Mapping_Issue": issue,
                }
                rows_out.append(base_row)
                mapped_rows += 1
                continue

            if sheet_name == "GST_Details":
                candidate_fsns, confidence, issue = resolve_join_fsn(
                    row,
                    columns,
                    target_fsns,
                    priority_map,
                    order_item_lookup,
                    order_id_lookup,
                    bridge_sku_lookup,
                    allow_ambiguous_sku_choice=False,
                    allow_sku_fallback=False,
                )
                if not candidate_fsns:
                    unmapped_rows += 1
                    continue
                fsn = candidate_fsns[0]
                fee_name = select_sheet_value(row, columns, ["fee_name"])
                target_field = map_gst_fee_name(fee_name)
                fee_amount = select_sheet_value(row, columns, ["fee_amount", "settlement_value"])
                gst_amount_total = parse_float(select_sheet_value(row, columns, ["cgst_amount"]))
                gst_amount_total += parse_float(select_sheet_value(row, columns, ["sgst_utgst_amount"]))
                gst_amount_total += parse_float(select_sheet_value(row, columns, ["igst_amount"]))
                gst_amount = format_decimal(gst_amount_total) if gst_amount_total else select_sheet_value(row, columns, ["fee_amount"])
                base_row = {
                    "FSN": fsn,
                    "Order_ID": select_sheet_value(row, columns, ["order_id"]),
                    "Order_Item_ID": select_sheet_value(row, columns, ["order_item_id_join", "order_item_id"]),
                    "Settlement_ID": select_sheet_value(row, columns, ["neft_id"]),
                    "Settlement_Date": select_sheet_value(row, columns, ["settlement_date", "payment_date"]),
                    "Seller_SKU": select_sheet_value(row, columns, ["sku_id"]),
                    "Gross_Amount": "",
                    "Commission": "",
                    "Fixed_Fee": "",
                    "Collection_Fee": "",
                    "Shipping_Fee": "",
                    "Reverse_Shipping_Fee": "",
                    "GST_On_Fees": gst_amount if target_field == "GST_On_Fees" else "",
                    "TCS": fee_amount if target_field == "TCS" else "",
                    "TDS": fee_amount if target_field == "TDS" else "",
                    "Refund": fee_amount if target_field == "Refund" else "",
                    "Protection_Fund": fee_amount if target_field == "Protection_Fund" else "",
                    "Adjustments": fee_amount if target_field == "Adjustments" else "",
                    "Net_Settlement": "",
                    "Source_File": file_path.name,
                    "Source_Sheet": sheet_name,
                    "Mapping_Confidence": confidence,
                    "Mapping_Issue": issue or ("Mapped through order_item_id" if select_sheet_value(row, columns, ["order_item_id"]) else ""),
                }
                if target_field in {"Commission", "Fixed_Fee", "Collection_Fee", "Shipping_Fee", "Reverse_Shipping_Fee"}:
                    base_row[target_field] = fee_amount
                elif target_field == "GST_On_Fees":
                    base_row["GST_On_Fees"] = gst_amount
                elif target_field == "TCS":
                    base_row["TCS"] = fee_amount
                elif target_field == "TDS":
                    base_row["TDS"] = fee_amount
                elif target_field == "Refund":
                    base_row["Refund"] = fee_amount
                elif target_field == "Protection_Fund":
                    base_row["Protection_Fund"] = fee_amount
                else:
                    base_row["Adjustments"] = fee_amount
                rows_out.append(base_row)
                mapped_rows += 1
                continue

            if sheet_name == "Non_Order_SPF":
                candidate_fsns, confidence, issue = resolve_join_fsn(
                    row,
                    columns,
                    target_fsns,
                    priority_map,
                    order_item_lookup,
                    order_id_lookup,
                    bridge_sku_lookup,
                    allow_ambiguous_sku_choice=False,
                    allow_sku_fallback=False,
                )
                if not candidate_fsns:
                    unmapped_rows += 1
                    continue
                settlement_value = select_sheet_value(row, columns, ["fee_amount", "settlement_value"])
                fsn = candidate_fsns[0]
                rows_out.append(
                    {
                        "FSN": fsn,
                        "Order_ID": select_sheet_value(row, columns, ["order_id"]),
                        "Order_Item_ID": select_sheet_value(row, columns, ["order_item_id_join", "order_item_id"]),
                        "Settlement_ID": select_sheet_value(row, columns, ["claim_id"]),
                        "Settlement_Date": select_sheet_value(row, columns, ["payment_date"]),
                        "Seller_SKU": select_sheet_value(row, columns, ["sku_id"]),
                        "Gross_Amount": "",
                        "Commission": "",
                        "Fixed_Fee": "",
                        "Collection_Fee": "",
                        "Shipping_Fee": "",
                        "Reverse_Shipping_Fee": "",
                        "GST_On_Fees": "",
                        "TCS": "",
                        "TDS": "",
                        "Refund": "",
                        "Protection_Fund": settlement_value,
                        "Adjustments": "",
                        "Net_Settlement": settlement_value,
                        "Source_File": file_path.name,
                        "Source_Sheet": sheet_name,
                        "Mapping_Confidence": confidence,
                        "Mapping_Issue": issue or "Direct FSN in source sheet",
                    }
                )
                mapped_rows += 1
                continue

            if sheet_name == "Storage_Recall":
                candidate_fsns, confidence, issue = resolve_join_fsn(
                    row,
                    columns,
                    target_fsns,
                    priority_map,
                    order_item_lookup,
                    order_id_lookup,
                    bridge_sku_lookup,
                    allow_ambiguous_sku_choice=False,
                    allow_sku_fallback=False,
                )
                if not candidate_fsns:
                    unmapped_rows += 1
                    continue
                fsn = candidate_fsns[0]
                marketplace_fees = select_sheet_value(row, columns, ["marketplace_fees", "fee_amount"])
                gst_fees = select_sheet_value(row, columns, ["gst_on_fees"])
                removal_fee = select_sheet_value(row, columns, ["removal_fee"])
                storage_fee = select_sheet_value(row, columns, ["storage_fee"])
                rows_out.append(
                    {
                        "FSN": fsn,
                        "Order_ID": "",
                        "Order_Item_ID": "",
                        "Settlement_ID": select_sheet_value(row, columns, ["recall_id", "listing_id", "neft_id"]),
                        "Settlement_Date": select_sheet_value(row, columns, ["payment_date"]),
                        "Seller_SKU": select_sheet_value(row, columns, ["sku_id"]),
                        "Gross_Amount": "",
                        "Commission": "",
                        "Fixed_Fee": marketplace_fees,
                        "Collection_Fee": "",
                        "Shipping_Fee": "",
                        "Reverse_Shipping_Fee": "",
                        "GST_On_Fees": gst_fees,
                        "TCS": "",
                        "TDS": "",
                        "Refund": "",
                        "Protection_Fund": "",
                        "Adjustments": first_non_blank(removal_fee, storage_fee),
                        "Net_Settlement": select_sheet_value(row, columns, ["settlement_value"]),
                        "Source_File": file_path.name,
                        "Source_Sheet": sheet_name,
                        "Mapping_Confidence": confidence,
                        "Mapping_Issue": issue or "Direct FSN in source sheet",
                    }
                )
                mapped_rows += 1
                continue

            unmapped_rows += 1

        total_raw_rows += raw_rows
        total_mapped_rows += mapped_rows
        total_unmapped_rows += unmapped_rows
        sheet_stats.append(
            {
                "sheet_name": sheet_name,
                "raw_rows": raw_rows,
                "mapped_rows": mapped_rows,
                "unmapped_rows": unmapped_rows,
                "status": "SUCCESS" if mapped_rows else "EMPTY",
                "columns": sorted(columns.keys()),
            }
        )

    return rows_out, {
        "sheets_scanned": len(relevant_sheets),
        "raw_rows": total_raw_rows,
        "mapped_rows": total_mapped_rows,
        "unmapped_rows": total_unmapped_rows,
        "sheet_stats": sheet_stats,
    }, selected_headers, sheet_stats


def parse_pnl_workbook(
    file_path: Path,
    target_fsns: Dict[str, Dict[str, Any]],
    order_item_lookup: Dict[str, str],
    order_id_lookup: Dict[str, str],
    bridge_sku_lookup: Dict[str, List[str]],
) -> Tuple[List[Dict[str, Any]], Dict[str, Any], List[Dict[str, Any]], List[Dict[str, Any]]]:
    workbook_rows = read_full_workbook_rows(file_path)
    rows_out: List[Dict[str, Any]] = []
    selected_headers: List[Dict[str, Any]] = []
    sheet_stats: List[Dict[str, Any]] = []
    total_raw_rows = 0
    total_mapped_rows = 0
    total_unmapped_rows = 0
    priority_map = {fsn: row for fsn, row in target_fsns.items()}
    relevant_sheets = ["Orders P&L", "SKU-level P&L"]

    for sheet_name in relevant_sheets:
        rows = workbook_rows.get(sheet_name, [])
        if not rows:
            sheet_stats.append({"sheet_name": sheet_name, "raw_rows": 0, "mapped_rows": 0, "unmapped_rows": 0, "status": "EMPTY"})
            continue

        header_rows = [0, 1]
        data_start = 2
        columns = build_columns_from_header_rows(rows, header_rows, load_synonyms())
        selected_headers.append(
            {
                "file_name": file_path.name,
                "sheet_name": sheet_name,
                "header_row_indexes": header_rows,
                "columns": sorted(columns.keys()),
                "header_preview": flatten_header_rows(rows, header_rows)[:15],
            }
        )

        data_rows = [row for row in rows[data_start:] if any(normalize_text(cell) for cell in row)]
        raw_rows = len(data_rows)
        mapped_rows = 0
        unmapped_rows = 0

        for row in data_rows:
            allow_sku_fallback = sheet_name == "SKU-level P&L"
            allow_ambiguous = False
            candidate_fsns, confidence, issue = resolve_join_fsn(
                row,
                columns,
                target_fsns,
                priority_map,
                order_item_lookup,
                order_id_lookup,
                bridge_sku_lookup,
                allow_ambiguous_sku_choice=allow_ambiguous,
                allow_sku_fallback=allow_sku_fallback,
            )
            if not candidate_fsns:
                unmapped_rows += 1
                continue

            fsn = candidate_fsns[0]
            row_out = {
                "FSN": fsn,
                "Order_ID": select_sheet_value(row, columns, ["order_id"]),
                "Order_Item_ID": select_sheet_value(row, columns, ["order_item_id"]),
                "Seller_SKU": select_sheet_value(row, columns, ["sku_id", "product_name"]),
                "Flipkart_Net_Earnings": select_sheet_value(row, columns, ["flipkart_net_earnings", "net_earnings"]),
                "Flipkart_Margin": select_sheet_value(row, columns, ["flipkart_margin", "margin"]),
                "Flipkart_Expenses": select_sheet_value(row, columns, ["flipkart_expenses", "total_expenses"]),
                "Amount_Settled": select_sheet_value(row, columns, ["amount_settled"]),
                "Amount_Pending": select_sheet_value(row, columns, ["amount_pending"]),
                "Source_File": file_path.name,
                "Source_Sheet": sheet_name,
                "Mapping_Confidence": confidence,
                "Mapping_Issue": issue,
            }
            if sheet_name == "SKU-level P&L" and not row_out["Order_ID"] and not row_out["Order_Item_ID"]:
                row_out["Mapping_Issue"] = issue or "Mapped via SKU bridge"
            rows_out.append(row_out)
            mapped_rows += 1

        total_raw_rows += raw_rows
        total_mapped_rows += mapped_rows
        total_unmapped_rows += unmapped_rows
        sheet_stats.append(
            {
                "sheet_name": sheet_name,
                "raw_rows": raw_rows,
                "mapped_rows": mapped_rows,
                "unmapped_rows": unmapped_rows,
                "status": "SUCCESS" if mapped_rows else "EMPTY",
                "columns": sorted(columns.keys()),
            }
        )

    return rows_out, {
        "sheets_scanned": len(relevant_sheets),
        "raw_rows": total_raw_rows,
        "mapped_rows": total_mapped_rows,
        "unmapped_rows": total_unmapped_rows,
        "sheet_stats": sheet_stats,
    }, selected_headers, sheet_stats


def normalize_flipkart_reports() -> Dict[str, Any]:
    ensure_directories()
    analysis = load_analysis()
    patterns = load_report_patterns()
    synonyms = load_synonyms()
    bridge_sku_lookup = load_bridge_sku_lookup()
    target_fsns = load_target_fsns()
    priority_map = {fsn: row for fsn, row in target_fsns.items()}
    target_fsn_set = set(target_fsns)

    outputs = {
        "listing": [],
        "orders": [],
        "returns": [],
        "settlements": [],
        "pnl": [],
        "sales_tax": [],
        "ads": [],
    }
    logs: List[Dict[str, Any]] = []
    order_lookup: Dict[str, str] = {}
    order_id_lookup: Dict[str, str] = {}
    report_stats: Dict[str, Dict[str, Any]] = {
        report_type: {
            "rows_before_filter": 0,
            "rows_after_fsn_filter": 0,
            "unmapped_rows": 0,
            "reason_if_zero": "",
            "candidate_sheets": [],
            "selected_sheet": {},
            "rejected_sheets": [],
        }
        for report_type in outputs
    }
    order_raw_fsns: List[str] = []
    sales_tax_unusable = False
    raw_files = list_input_files(RAW_INPUT_DIR)
    for report_type in REPORT_TYPES:
        selection = select_best_sheet_across_files(raw_files, report_type, synonyms, patterns)
        report_stats[report_type]["candidate_sheets"].extend(selection.get("candidates", []))
        report_stats[report_type]["rejected_sheets"].extend(selection.get("rejected_sheets", []))
        print_selection(report_type, selection)

        selected = selection.get("selected_sheet")
        if not selected:
            reason_if_zero = "No usable settlement transaction detail sheet found in uploaded file." if report_type == "settlements" else "No usable data sheet found in uploaded file."
            if report_type == "settlements":
                print(reason_if_zero)
            report_stats[report_type]["reason_if_zero"] = report_stats[report_type]["reason_if_zero"] or reason_if_zero
            logs.append(
                {
                    "timestamp": now_iso(),
                    "output_file": report_type,
                    "source_file": "",
                    "sheet_name": "",
                    "rows_read": 0,
                    "rows_written": 0,
                    "rows_before_filter": 0,
                    "rows_after_fsn_filter": 0,
                    "unmapped_rows": 0,
                    "reason_if_zero": reason_if_zero,
                    "status": "EMPTY",
                    "message": "no usable sheet",
                }
            )
            continue

        file_name = selected.get("file_name", "")
        file_path = Path(selected.get("file_path", RAW_INPUT_DIR / file_name))
        sheet_name = selected.get("sheet_name", "")
        workbook_rows = read_workbook_rows(file_path)
        rows = workbook_rows.get(sheet_name, [])
        header_row_index = int(selected.get("header_row_index", 0))
        columns = selected.get("detected_columns", {})
        data_rows = rows[header_row_index + 1 :] if rows else []
        rows_before_filter = len(data_rows)
        rows_after_filter = 0
        unmapped_rows = 0
        join_key_detected = extract_join_key_notes(columns)

        report_stats[report_type]["selected_sheet"] = {
            "file_name": file_name,
            "sheet_name": sheet_name,
            "header_row_index": header_row_index,
            "selection_score": selected.get("selection_score", 0.0),
            "combined_score": selected.get("combined_score", 0.0),
            "columns": sorted(columns.keys()),
            "raw_row_count": selected.get("raw_row_count", rows_before_filter),
            "normalized_row_count": selected.get("non_empty_row_count", rows_before_filter),
        }
        report_stats[report_type]["rows_before_filter"] += rows_before_filter

        if report_type == "orders":
            order_raw_fsns.extend([clean_fsn(value(row, columns, "fsn")) for row in data_rows if clean_fsn(value(row, columns, "fsn"))])
        if report_type == "pnl" and selected.get("useful_column_count", 0) == 0:
            report_stats[report_type]["reason_if_zero"] = "No usable P&L detail sheet found in uploaded file."

        if report_type == "sales_tax" and not join_key_detected:
            sales_tax_unusable = True
            report_stats[report_type]["reason_if_zero"] = "Not usable for FSN-level analysis from current file structure"
            logs.append(
                {
                    "timestamp": now_iso(),
                    "output_file": report_type,
                    "source_file": file_name,
                    "sheet_name": sheet_name,
                    "rows_read": rows_before_filter,
                    "rows_written": 0,
                    "rows_before_filter": rows_before_filter,
                    "rows_after_fsn_filter": 0,
                    "unmapped_rows": rows_before_filter,
                    "reason_if_zero": report_stats[report_type]["reason_if_zero"],
                    "status": "EMPTY",
                    "message": "join key missing",
                }
            )
            continue

        if rows_before_filter == 0:
            reason_if_zero = build_reason(report_type, 0, 0, 0, columns, join_key_detected)
            if report_type == "settlements":
                reason_if_zero = "No usable settlement transaction detail sheet found in uploaded file."
            report_stats[report_type]["reason_if_zero"] = report_stats[report_type]["reason_if_zero"] or reason_if_zero
            logs.append(
                {
                    "timestamp": now_iso(),
                    "output_file": report_type,
                    "source_file": file_name,
                    "sheet_name": sheet_name,
                    "rows_read": 0,
                    "rows_written": 0,
                    "rows_before_filter": 0,
                    "rows_after_fsn_filter": 0,
                    "unmapped_rows": 0,
                    "reason_if_zero": reason_if_zero,
                    "status": "EMPTY",
                    "message": "no rows after header",
                }
            )
            continue

        for row in data_rows:
            candidate_fsns, confidence, issue = resolve_fsn_from_row(
                row,
                columns,
                target_fsns,
                priority_map,
                order_lookup,
                order_id_lookup,
                bridge_sku_lookup,
                allow_ambiguous_sku_choice=report_type != "pnl",
            )
            if not candidate_fsns:
                unmapped_rows += 1
                continue

            for fsn in candidate_fsns:
                fsn = clean_fsn(fsn)
                if not fsn or fsn not in target_fsn_set:
                    continue
                rows_after_filter += 1

                if report_type == "listing":
                    outputs["listing"].append(
                        {
                            "FSN": fsn,
                            "Seller_SKU": value(row, columns, "sku_id"),
                            "Product_Title": value(row, columns, "product_title"),
                            "Category": value(row, columns, "category"),
                            "Listing_Status": value(row, columns, "listing_status"),
                            "Inactive_Reason": value(row, columns, "inactive_reason"),
                            "MRP": value(row, columns, "mrp"),
                            "Selling_Price": value(row, columns, "selling_price"),
                            "Stock": value(row, columns, "stock"),
                            "Bank_Settlement": value(row, columns, "bank_settlement"),
                            "Listing_Quality": value(row, columns, "listing_quality"),
                            "Package_Length": value(row, columns, "package_length"),
                            "Package_Breadth": value(row, columns, "package_breadth"),
                            "Package_Height": value(row, columns, "package_height"),
                            "Package_Weight": value(row, columns, "package_weight"),
                            "HSN": value(row, columns, "hsn"),
                            "Tax_Code": value(row, columns, "tax_code"),
                            "Source_File": file_name,
                            "Mapping_Confidence": confidence,
                        }
                    )
                elif report_type == "orders":
                    order_id = clean_fsn(value(row, columns, "order_id"))
                    order_item_id = clean_fsn(value(row, columns, "order_item_id"))
                    outputs["orders"].append(
                        {
                            "FSN": fsn,
                            "Order_ID": order_id,
                            "Order_Item_ID": order_item_id,
                            "Seller_SKU": value(row, columns, "sku_id"),
                            "Product_Title": value(row, columns, "product_title"),
                            "Order_Date": value(row, columns, "order_date"),
                            "Quantity": value(row, columns, "quantity"),
                            "Selling_Price": value(row, columns, "selling_price"),
                            "Order_Status": value(row, columns, "order_status"),
                            "Dispatch_Date": value(row, columns, "dispatch_date"),
                            "Delivery_Date": value(row, columns, "delivery_date"),
                            "Cancellation_Status": value(row, columns, "cancellation_status"),
                            "Source_File": file_name,
                            "Mapping_Confidence": confidence,
                        }
                    )
                    if order_item_id:
                        for variant in order_identifier_variants(order_item_id):
                            order_lookup.setdefault(variant, fsn)
                    if order_id:
                        for variant in order_identifier_variants(order_id):
                            order_id_lookup.setdefault(variant, fsn)
                elif report_type == "returns":
                    outputs["returns"].append(
                        {
                            "FSN": fsn,
                            "Order_ID": clean_fsn(value(row, columns, "order_id")),
                            "Order_Item_ID": clean_fsn(value(row, columns, "order_item_id")),
                            "Seller_SKU": value(row, columns, "sku_id"),
                            "Return_ID": clean_fsn(value(row, columns, "return_id")),
                            "Return_Date": value(row, columns, "return_date"),
                            "Return_Type": value(row, columns, "return_type"),
                            "Return_Reason": value(row, columns, "return_reason"),
                            "Return_Status": value(row, columns, "return_status"),
                            "Reverse_Shipment_Status": value(row, columns, "reverse_shipment_status"),
                            "Source_File": file_name,
                            "Mapping_Confidence": confidence,
                        }
                    )
                elif report_type == "settlements":
                    outputs["settlements"].append(
                        {
                            "FSN": fsn,
                            "Order_ID": clean_fsn(value(row, columns, "order_id")),
                            "Order_Item_ID": clean_fsn(value(row, columns, "order_item_id")),
                            "Settlement_ID": clean_fsn(value(row, columns, "settlement_id")),
                            "Settlement_Date": value(row, columns, "settlement_date"),
                            "Seller_SKU": value(row, columns, "sku_id"),
                            "Gross_Amount": value(row, columns, "gross_amount"),
                            "Commission": value(row, columns, "commission"),
                            "Fixed_Fee": value(row, columns, "fixed_fee"),
                            "Collection_Fee": value(row, columns, "collection_fee"),
                            "Shipping_Fee": value(row, columns, "shipping_fee"),
                            "Reverse_Shipping_Fee": value(row, columns, "reverse_shipping_fee"),
                            "GST_On_Fees": value(row, columns, "gst_on_fees"),
                            "TCS": value(row, columns, "tcs"),
                            "TDS": value(row, columns, "tds"),
                            "Refund": value(row, columns, "refund"),
                            "Protection_Fund": value(row, columns, "protection_fund"),
                            "Adjustments": value(row, columns, "adjustments"),
                            "Net_Settlement": value(row, columns, "net_settlement"),
                            "Source_File": file_name,
                            "Mapping_Confidence": confidence,
                        }
                    )
                elif report_type == "pnl":
                    outputs["pnl"].append(
                        {
                            "FSN": fsn,
                            "Order_ID": clean_fsn(value(row, columns, "order_id")),
                            "Order_Item_ID": clean_fsn(value(row, columns, "order_item_id")),
                            "Seller_SKU": value(row, columns, "sku_id"),
                            "Flipkart_Net_Earnings": value(row, columns, "flipkart_net_earnings"),
                            "Flipkart_Margin": value(row, columns, "flipkart_margin"),
                            "Flipkart_Expenses": value(row, columns, "flipkart_expenses"),
                            "Amount_Settled": value(row, columns, "amount_settled"),
                            "Amount_Pending": value(row, columns, "amount_pending"),
                            "Source_File": file_name,
                            "Mapping_Confidence": confidence,
                        }
                    )
                elif report_type == "sales_tax":
                    outputs["sales_tax"].append(
                        {
                            "FSN": fsn,
                            "Order_ID": clean_fsn(value(row, columns, "order_id")),
                            "Order_Item_ID": clean_fsn(value(row, columns, "order_item_id")),
                            "Invoice_ID": value(row, columns, "invoice_id"),
                            "Invoice_Date": value(row, columns, "invoice_date"),
                            "HSN": value(row, columns, "hsn"),
                            "Taxable_Value": value(row, columns, "taxable_value"),
                            "IGST": value(row, columns, "igst"),
                            "CGST": value(row, columns, "cgst"),
                            "SGST": value(row, columns, "sgst"),
                            "TCS": value(row, columns, "tcs"),
                            "TDS": value(row, columns, "tds"),
                            "Event_Type": value(row, columns, "event_type"),
                            "Source_File": file_name,
                            "Mapping_Confidence": confidence,
                        }
                    )
                elif report_type == "ads":
                    sku = value(row, columns, "sku_id")
                    ad_fsns: List[str] = []
                    fsn_value = clean_fsn(value(row, columns, "fsn"))
                    if fsn_value and fsn_value in target_fsn_set:
                        ad_fsns = [fsn_value]
                    elif sku:
                        ad_fsns = [clean_fsn(fsn_id) for fsn_id, target_row in target_fsns.items() if normalize_text(target_row.get("SKU_ID")) == sku]
                    if not ad_fsns:
                        continue
                    roi = parse_float(value(row, columns, "roi"))
                    total_revenue = parse_float(value(row, columns, "total_revenue"))
                    estimated_ad_spend = total_revenue / roi if roi > 0 else 0.0
                    selected_fsn = clean_fsn(ad_fsns[0]) if len(ad_fsns) == 1 else clean_fsn(highest_priority_fsn(ad_fsns, priority_map))
                    if not selected_fsn:
                        continue
                    outputs["ads"].append(
                        {
                            "FSN": selected_fsn,
                            "Campaign_ID": value(row, columns, "campaign_id"),
                            "Campaign_Name": value(row, columns, "campaign_name"),
                            "AdGroup_ID": value(row, columns, "adgroup_id"),
                            "AdGroup_Name": value(row, columns, "adgroup_name"),
                            "Seller_SKU": sku,
                            "Product_Name": value(row, columns, "product_name"),
                            "Views": value(row, columns, "views"),
                            "Clicks": value(row, columns, "clicks"),
                            "Direct_Units_Sold": value(row, columns, "direct_units_sold"),
                            "Indirect_Units_Sold": value(row, columns, "indirect_units_sold"),
                            "Total_Revenue": value(row, columns, "total_revenue"),
                            "ROI": value(row, columns, "roi"),
                            "Estimated_Ad_Spend": f"{estimated_ad_spend:.2f}" if estimated_ad_spend else "",
                            "ROAS": value(row, columns, "roi"),
                            "ACOS": f"{(estimated_ad_spend / total_revenue):.4f}" if total_revenue > 0 and estimated_ad_spend else "",
                            "Source_File": file_name,
                            "Mapping_Confidence": "LOW" if len(ad_fsns) > 1 else confidence,
                            "Mapping_Issue": "SKU maps to multiple FSNs" if len(ad_fsns) > 1 else "",
                        }
                    )

        if report_type == "orders":
            intersection = len(set(order_raw_fsns) & target_fsn_set)
            print(f"Selected order sheet: {sheet_name}")
            print(f"Selected order header row: {header_row_index}")
            print(f"Selected order columns: {sorted(columns.keys())}")
            print(f"Selected order raw rows: {rows_before_filter}")
            print(f"First 5 real FSNs: {order_raw_fsns[:5]}")
            print(f"First 5 target FSNs: {sorted(target_fsn_set)[:5]}")
            print(f"Target FSN intersection count: {intersection}")
            if intersection == 0:
                report_stats["orders"]["reason_if_zero"] = "Order report FSNs do not overlap target FSN file."

        report_stats[report_type]["rows_after_fsn_filter"] += rows_after_filter

        logs.append(
            {
                "timestamp": now_iso(),
                "output_file": report_type,
                "source_file": file_name,
                "sheet_name": sheet_name,
                "rows_read": rows_before_filter,
                "rows_written": rows_after_filter,
                "rows_before_filter": rows_before_filter,
                "rows_after_fsn_filter": rows_after_filter,
                "unmapped_rows": unmapped_rows,
                "reason_if_zero": "",
                "status": "SUCCESS" if rows_after_filter else "EMPTY",
                "message": "normalized",
            }
        )

    order_rows = dedupe_dict_rows(outputs["orders"], "Order_Item_ID")
    if not order_rows:
        order_rows = dedupe_dict_rows(outputs["orders"], "Order_ID")
    special_order_item_lookup, special_order_id_lookup = build_order_lookup(order_rows)

    settlement_file = RAW_INPUT_DIR / SETTLEMENT_WORKBOOK_NAME
    settlement_rows: List[Dict[str, Any]] = []
    settlement_summary: Dict[str, Any] = {"sheets_scanned": 0, "raw_rows": 0, "mapped_rows": 0, "unmapped_rows": 0, "sheet_stats": []}
    settlement_headers: List[Dict[str, Any]] = []
    settlement_sheet_stats: List[Dict[str, Any]] = []
    if settlement_file.exists():
        settlement_rows, settlement_summary, settlement_headers, settlement_sheet_stats = parse_settlement_workbook(
            settlement_file,
            target_fsns,
            special_order_item_lookup,
            special_order_id_lookup,
            bridge_sku_lookup,
        )
        outputs["settlements"] = settlement_rows
        logs.append(
            {
                "timestamp": now_iso(),
                "output_file": "settlements",
                "source_file": settlement_file.name,
                "sheet_name": "multiple",
                "rows_read": settlement_summary["raw_rows"],
                "rows_written": settlement_summary["mapped_rows"],
                "rows_before_filter": settlement_summary["raw_rows"],
                "rows_after_fsn_filter": settlement_summary["mapped_rows"],
                "unmapped_rows": settlement_summary["unmapped_rows"],
                "reason_if_zero": "" if settlement_summary["mapped_rows"] else "No usable settlement transaction detail sheet found in uploaded file.",
                "status": "SUCCESS" if settlement_summary["mapped_rows"] else "EMPTY",
                "message": "special workbook parser",
            }
        )
        report_stats["settlements"]["rows_before_filter"] = settlement_summary["raw_rows"]
        report_stats["settlements"]["rows_after_fsn_filter"] = settlement_summary["mapped_rows"]
        report_stats["settlements"]["unmapped_rows"] = settlement_summary["unmapped_rows"]
        report_stats["settlements"]["reason_if_zero"] = "" if settlement_summary["mapped_rows"] else "No usable settlement transaction detail sheet found in uploaded file."
        report_stats["settlements"]["candidate_sheets"] = settlement_sheet_stats
        report_stats["settlements"]["selected_sheet"] = {
            "file_name": settlement_file.name,
            "sheet_name": "multiple",
            "header_row_index": 1,
            "selection_score": float(settlement_summary["mapped_rows"]),
            "combined_score": float(settlement_summary["mapped_rows"]),
            "columns": sorted({column for sheet in settlement_sheet_stats for column in sheet.get("columns", [])}),
            "raw_row_count": settlement_summary["raw_rows"],
            "normalized_row_count": settlement_summary["mapped_rows"],
        }

    pnl_file = RAW_INPUT_DIR / PNL_WORKBOOK_NAME
    pnl_rows: List[Dict[str, Any]] = []
    pnl_summary: Dict[str, Any] = {"sheets_scanned": 0, "raw_rows": 0, "mapped_rows": 0, "unmapped_rows": 0, "sheet_stats": []}
    pnl_headers: List[Dict[str, Any]] = []
    pnl_sheet_stats: List[Dict[str, Any]] = []
    if pnl_file.exists():
        pnl_rows, pnl_summary, pnl_headers, pnl_sheet_stats = parse_pnl_workbook(
            pnl_file,
            target_fsns,
            special_order_item_lookup,
            special_order_id_lookup,
            bridge_sku_lookup,
        )
        outputs["pnl"] = pnl_rows
        logs.append(
            {
                "timestamp": now_iso(),
                "output_file": "pnl",
                "source_file": pnl_file.name,
                "sheet_name": "Orders P&L + SKU-level P&L",
                "rows_read": pnl_summary["raw_rows"],
                "rows_written": pnl_summary["mapped_rows"],
                "rows_before_filter": pnl_summary["raw_rows"],
                "rows_after_fsn_filter": pnl_summary["mapped_rows"],
                "unmapped_rows": pnl_summary["unmapped_rows"],
                "reason_if_zero": "" if pnl_summary["mapped_rows"] else "No usable P&L detail sheet found in uploaded file.",
                "status": "SUCCESS" if pnl_summary["mapped_rows"] else "EMPTY",
                "message": "special workbook parser",
            }
        )
        report_stats["pnl"]["rows_before_filter"] = pnl_summary["raw_rows"]
        report_stats["pnl"]["rows_after_fsn_filter"] = pnl_summary["mapped_rows"]
        report_stats["pnl"]["unmapped_rows"] = pnl_summary["unmapped_rows"]
        report_stats["pnl"]["reason_if_zero"] = "" if pnl_summary["mapped_rows"] else "No usable P&L detail sheet found in uploaded file."
        report_stats["pnl"]["candidate_sheets"] = pnl_sheet_stats
        report_stats["pnl"]["selected_sheet"] = {
            "file_name": pnl_file.name,
            "sheet_name": "Orders P&L + SKU-level P&L",
            "header_row_index": 1,
            "selection_score": float(pnl_summary["mapped_rows"]),
            "combined_score": float(pnl_summary["mapped_rows"]),
            "columns": sorted({column for sheet in pnl_sheet_stats for column in sheet.get("columns", [])}),
            "raw_row_count": pnl_summary["raw_rows"],
            "normalized_row_count": pnl_summary["mapped_rows"],
        }

    for report_type, stats in report_stats.items():
        if stats["rows_after_fsn_filter"] > 0:
            stats["reason_if_zero"] = ""
            continue
        if report_type == "orders":
            stats["reason_if_zero"] = "Order report FSNs do not overlap target FSN file."
        elif report_type == "returns":
            stats["reason_if_zero"] = "normalized_orders is empty, so return mapping cannot run yet"
        elif report_type == "sales_tax" and sales_tax_unusable:
            stats["reason_if_zero"] = "Not usable for FSN-level analysis from current file structure"
        elif report_type == "settlements":
            stats["reason_if_zero"] = stats["reason_if_zero"] or "No usable settlement transaction detail sheet found in uploaded file."
        elif report_type == "pnl":
            stats["reason_if_zero"] = stats["reason_if_zero"] or "No usable P&L detail sheet found in uploaded file."
        elif not stats["reason_if_zero"]:
            stats["reason_if_zero"] = "FSNs were found, but none matched the target FSN list"

    write_csv(NORMALIZED_LISTINGS_PATH, LISTING_HEADERS, outputs["listing"])
    write_csv(NORMALIZED_ORDERS_PATH, ORDER_HEADERS, order_rows)
    write_csv(NORMALIZED_RETURNS_PATH, RETURN_HEADERS, outputs["returns"])
    write_csv(NORMALIZED_SETTLEMENTS_PATH, SETTLEMENT_HEADERS, outputs["settlements"])
    write_csv(NORMALIZED_PNL_PATH, PNL_HEADERS, outputs["pnl"])
    write_csv(NORMALIZED_SALES_TAX_PATH, SALES_TAX_HEADERS, outputs["sales_tax"])
    write_csv(NORMALIZED_ADS_PATH, ADS_HEADERS, outputs["ads"])
    append_csv_log(NORMALIZATION_LOG_PATH, LOG_HEADERS, logs)

    result = {
        "status": "SUCCESS",
        "generated_at": now_iso(),
        "analysis_path": str(ANALYSIS_JSON_PATH),
        "bridge_path": str(FSN_BRIDGE_PATH),
        "outputs": {
            "normalized_listings": len(outputs["listing"]),
            "normalized_orders": len(order_rows),
            "normalized_returns": len(outputs["returns"]),
            "normalized_settlements": len(outputs["settlements"]),
            "normalized_pnl": len(outputs["pnl"]),
            "normalized_sales_tax": len(outputs["sales_tax"]),
            "normalized_ads": len(outputs["ads"]),
        },
        "report_diagnostics": report_stats,
        "log_path": str(NORMALIZATION_LOG_PATH),
        "diagnostics": {
            "settlement_sheets_scanned": settlement_summary["sheets_scanned"],
            "settlement_raw_rows_found": settlement_summary["raw_rows"],
            "settlement_rows_mapped_to_fsn": settlement_summary["mapped_rows"],
            "settlement_unmapped_rows": settlement_summary["unmapped_rows"],
            "pnl_sheets_scanned": pnl_summary["sheets_scanned"],
            "pnl_raw_rows_found": pnl_summary["raw_rows"],
            "pnl_rows_mapped_to_fsn": pnl_summary["mapped_rows"],
            "pnl_unmapped_rows": pnl_summary["unmapped_rows"],
            "sample_selected_headers": {
                "settlements": settlement_headers[:3],
                "pnl": pnl_headers[:3],
            },
            "sample_first_3_mapped_rows": {
                "settlements": settlement_rows[:3],
                "pnl": pnl_rows[:3],
            },
        },
    }
    save_run_state(
        NORMALIZATION_STATE_PATH,
        {
            "status": "SUCCESS",
            "stage": "normalize",
            "generated_at": result["generated_at"],
            "source_analysis_mtime": ANALYSIS_JSON_PATH.stat().st_mtime,
            "source_bridge_mtime": FSN_BRIDGE_PATH.stat().st_mtime,
            "report_diagnostics": report_stats,
            "output_files": {
                "normalized_listings": {
                    "path": str(NORMALIZED_LISTINGS_PATH),
                    "rows": len(outputs["listing"]),
                    "mtime": NORMALIZED_LISTINGS_PATH.stat().st_mtime,
                },
                "normalized_orders": {
                    "path": str(NORMALIZED_ORDERS_PATH),
                    "rows": len(order_rows),
                    "mtime": NORMALIZED_ORDERS_PATH.stat().st_mtime,
                },
                "normalized_returns": {
                    "path": str(NORMALIZED_RETURNS_PATH),
                    "rows": len(outputs["returns"]),
                    "mtime": NORMALIZED_RETURNS_PATH.stat().st_mtime,
                },
                "normalized_settlements": {
                    "path": str(NORMALIZED_SETTLEMENTS_PATH),
                    "rows": len(outputs["settlements"]),
                    "mtime": NORMALIZED_SETTLEMENTS_PATH.stat().st_mtime,
                },
                "normalized_pnl": {
                    "path": str(NORMALIZED_PNL_PATH),
                    "rows": len(outputs["pnl"]),
                    "mtime": NORMALIZED_PNL_PATH.stat().st_mtime,
                },
                "normalized_sales_tax": {
                    "path": str(NORMALIZED_SALES_TAX_PATH),
                    "rows": len(outputs["sales_tax"]),
                    "mtime": NORMALIZED_SALES_TAX_PATH.stat().st_mtime,
                },
                "normalized_ads": {
                    "path": str(NORMALIZED_ADS_PATH),
                    "rows": len(outputs["ads"]),
                    "mtime": NORMALIZED_ADS_PATH.stat().st_mtime,
                },
            },
        },
    )
    payload = dict(result)
    payload.pop("status", None)
    print(json.dumps(build_status_payload("SUCCESS", **payload), indent=2))
    return result


def main() -> None:
    try:
        normalize_flipkart_reports()
    except Exception as exc:
        print(
            json.dumps(
                {
                    "status": "ERROR",
                    "error_type": exc.__class__.__name__,
                    "message": str(exc),
                    "outputs": {
                        "normalized_listings": str(NORMALIZED_LISTINGS_PATH),
                        "normalized_orders": str(NORMALIZED_ORDERS_PATH),
                        "normalized_returns": str(NORMALIZED_RETURNS_PATH),
                        "normalized_settlements": str(NORMALIZED_SETTLEMENTS_PATH),
                        "normalized_pnl": str(NORMALIZED_PNL_PATH),
                        "normalized_sales_tax": str(NORMALIZED_SALES_TAX_PATH),
                        "normalized_ads": str(NORMALIZED_ADS_PATH),
                    },
                    "log_path": str(NORMALIZATION_LOG_PATH),
                },
                indent=2,
            )
        )
        raise SystemExit(1)


if __name__ == "__main__":
    main()

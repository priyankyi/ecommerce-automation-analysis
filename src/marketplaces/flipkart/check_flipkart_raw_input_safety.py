from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.marketplaces.flipkart.flipkart_utils import (  # noqa: E402
    LOG_DIR,
    OUTPUT_DIR,
    RAW_INPUT_DIR,
    append_csv_log,
    file_mtime_iso,
    list_input_files,
    infer_report_type as infer_report_type_scored,
    load_report_patterns,
    load_synonyms,
    normalize_key,
    normalize_text,
    now_iso,
    write_csv,
)
from src.marketplaces.flipkart.report_format_monitor_utils import (  # noqa: E402
    detect_header_row,
)

RAW_MANIFEST_PATH = OUTPUT_DIR / "flipkart_raw_input_manifest.csv"
LATEST_MANIFEST_PATH = OUTPUT_DIR / "flipkart_latest_raw_input_manifest.json"
LOG_PATH = LOG_DIR / "flipkart_raw_input_safety_log.csv"

MANIFEST_HEADERS = [
    "file_name",
    "file_path",
    "file_size",
    "modified_time",
    "sha256_hash",
    "detected_report_type",
    "detected_period",
    "warning",
]

MONTH_NAME_MAP = {
    "jan": "January",
    "january": "January",
    "feb": "February",
    "february": "February",
    "mar": "March",
    "march": "March",
    "apr": "April",
    "april": "April",
    "may": "May",
    "jun": "June",
    "june": "June",
    "jul": "July",
    "july": "July",
    "aug": "August",
    "august": "August",
    "sep": "September",
    "sept": "September",
    "september": "September",
    "oct": "October",
    "october": "October",
    "nov": "November",
    "november": "November",
    "dec": "December",
    "december": "December",
}


def _ensure_output_dirs() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)


def _json_text(payload: Dict[str, Any]) -> str:
    return json.dumps(payload, indent=2, ensure_ascii=False)


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _preview_csv_rows(path: Path, max_rows: int = 20) -> List[List[Any]]:
    rows: List[List[Any]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            rows.append(list(row))
            if len(rows) >= max_rows:
                break
    return rows


def _preview_xlsx_rows(path: Path, max_rows: int = 20) -> Dict[str, List[List[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency failure
        raise RuntimeError("Missing dependency `openpyxl`. Install it with: python -m pip install openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    preview: Dict[str, List[List[Any]]] = {}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows: List[List[Any]] = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) >= max_rows:
                break
        preview[sheet_name] = rows
    return preview


def _preview_xls_rows(path: Path, max_rows: int = 20) -> Dict[str, List[List[Any]]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - dependency failure
        raise RuntimeError("Missing dependency `xlrd`. Install it with: python -m pip install xlrd") from exc

    workbook = xlrd.open_workbook(str(path))
    preview: Dict[str, List[List[Any]]] = {}
    for sheet_name in workbook.sheet_names():
        sheet = workbook.sheet_by_name(sheet_name)
        rows: List[List[Any]] = []
        for row_index in range(min(sheet.nrows, max_rows)):
            rows.append(sheet.row_values(row_index))
        preview[sheet_name] = rows
    return preview


def _load_preview_rows(path: Path, max_rows: int = 20) -> Dict[str, List[List[Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return {path.name: _preview_csv_rows(path, max_rows=max_rows)}
    if suffix in {".xlsx", ".xlsm"}:
        return _preview_xlsx_rows(path, max_rows=max_rows)
    if suffix == ".xls":
        return _preview_xls_rows(path, max_rows=max_rows)
    return {path.name: []}


def _join_text(*parts: Any) -> str:
    return " ".join(normalize_key(part) for part in parts if normalize_text(part))


def _category_from_type(
    report_type: str,
    file_name: str,
    sheet_name: str,
    headers: Sequence[Any],
    detected_columns: Dict[str, Dict[str, Any]],
) -> str:
    filename_text = _join_text(file_name, sheet_name)
    header_keys = set(detected_columns)

    if "listing status" in filename_text or "listingstatus" in filename_text:
        return "listing_status"
    if report_type in {"pnl", "sales_tax"} or "pnl" in filename_text or "profit" in filename_text or "earnings" in filename_text or "sales report" in filename_text:
        return "payment"
    if report_type == "settlements" or "settlement" in filename_text or "settled" in filename_text:
        return "settlement"
    if report_type == "orders" or "order" in filename_text or {"order_id", "order_item_id"} & header_keys:
        return "orders"
    if report_type == "returns" or "return" in filename_text:
        return "returns"
    if report_type == "ads" or "ads" in filename_text or "campaign" in filename_text:
        return "ads"
    if report_type == "master_fsn" or "master fsn" in filename_text or "masterfsn" in filename_text:
        return "inventory"
    if report_type == "listing":
        if "listing status" in filename_text or {"listing_status", "inactive_reason"} & header_keys:
            return "listing_status"
        if {"stock", "quantity", "available", "inventory"} & header_keys:
            return "inventory"
        return "listing"
    if "listing status" in filename_text or {"listing_status", "inactive_reason"} & header_keys:
        return "listing_status"
    if "listing" in filename_text or "catalog" in filename_text:
        return "listing"
    if "inventory" in filename_text or {"stock", "quantity", "available"} & header_keys:
        return "inventory"
    return "unknown"


def _detect_period_from_filename(file_name: str) -> str:
    text = normalize_text(file_name)
    if not text:
        return "Unknown"

    year_month_match = re.search(r"(?<!\d)((?:19|20)\d{2})[-_/](0?[1-9]|1[0-2])(?!\d)", text)
    if year_month_match:
        year = year_month_match.group(1)
        month = int(year_month_match.group(2))
        return f"{year}-{month:02d}"

    month_year_match = re.search(
        r"(?i)\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|"
        r"sep(?:t|tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\s+((?:19|20)\d{2})\b",
        text,
    )
    if month_year_match:
        month_token = month_year_match.group(1).lower()
        return f"{MONTH_NAME_MAP.get(month_token, month_token.title())} {month_year_match.group(2)}"

    ranged_match = re.search(
        r"(?i)\b((?:\d{4}[-_/]\d{1,2}[-_/]\d{1,2})|(?:\d{1,2}[-_/]\d{1,2}[-_/]\d{4})|(?:\d{1,2}[A-Za-z]{3}\d{4}))"
        r"\s*(?:to|[-_])\s*"
        r"((?:\d{4}[-_/]\d{1,2}[-_/]\d{1,2})|(?:\d{1,2}[-_/]\d{1,2}[-_/]\d{4})|(?:\d{1,2}[A-Za-z]{3}\d{4}))\b",
        text,
    )
    if ranged_match:
        start = ranged_match.group(1).replace("_", "-")
        end = ranged_match.group(2).replace("_", "-")
        return f"{start} to {end}"

    return "Unknown"


def _normalize_signature_rows(rows: Sequence[Dict[str, Any]]) -> List[str]:
    signatures = [
        "|".join(
            [
                normalize_text(row.get("file_name", "")),
                normalize_text(row.get("file_size", "")),
                normalize_text(row.get("modified_time", "")),
                normalize_text(row.get("detected_report_type", "")),
                normalize_text(row.get("sha256_hash", "")),
            ]
        )
        for row in rows
    ]
    return sorted(signature for signature in signatures if signature)


def _normalize_signature_rows_without_hash(rows: Sequence[Dict[str, Any]]) -> List[str]:
    signatures = [
        "|".join(
            [
                normalize_text(row.get("file_name", "")),
                normalize_text(row.get("file_size", "")),
                normalize_text(row.get("modified_time", "")),
                normalize_text(row.get("detected_report_type", "")),
            ]
        )
        for row in rows
    ]
    return sorted(signature for signature in signatures if signature)


def _normalize_identity_rows(rows: Sequence[Dict[str, Any]]) -> List[str]:
    signatures = [
        "|".join(
            [
                normalize_text(row.get("file_name", "")),
                normalize_text(row.get("file_size", "")),
                normalize_text(row.get("modified_time", "")),
            ]
        )
        for row in rows
    ]
    return sorted(signature for signature in signatures if signature)


def _find_latest_run_manifest_path() -> Path | None:
    runs_dir = OUTPUT_DIR / "runs"
    if not runs_dir.exists():
        return None

    latest_path: Path | None = None
    latest_mtime = -1.0
    for run_dir in runs_dir.iterdir():
        if not run_dir.is_dir() or not run_dir.name.startswith("FLIPKART_"):
            continue
        summary_path = run_dir / "pipeline_run_summary.json"
        manifest_path = run_dir / "input_manifest.csv"
        if not summary_path.exists() or not manifest_path.exists():
            continue
        candidate_mtime = summary_path.stat().st_mtime
        if candidate_mtime > latest_mtime:
            latest_mtime = candidate_mtime
            latest_path = manifest_path
    return latest_path


def _read_manifest_rows(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def _current_file_signature(row: Dict[str, Any]) -> str:
    return "|".join(
        [
            normalize_text(row.get("file_name", "")),
            normalize_text(row.get("file_size", "")),
            normalize_text(row.get("modified_time", "")),
            normalize_text(row.get("detected_report_type", "")),
        ]
    )


def _row_warning_text(warnings: Iterable[str]) -> str:
    unique_warnings = [warning for warning in dict.fromkeys(normalize_text(item) for item in warnings if normalize_text(item))]
    return " | ".join(unique_warnings)


def _write_latest_manifest_json(payload: Dict[str, Any]) -> None:
    LATEST_MANIFEST_PATH.write_text(_json_text(payload), encoding="utf-8")


def _append_log(result: Dict[str, Any]) -> None:
    append_csv_log(
        LOG_PATH,
        [
            "timestamp",
            "status",
            "raw_file_count",
            "duplicate_file_count",
            "unknown_file_count",
            "safe_to_run_full_refresh",
            "manifest_path",
            "latest_manifest_path",
            "message",
        ],
        [
            {
                "timestamp": now_iso(),
                "status": result.get("status", ""),
                "raw_file_count": result.get("raw_file_count", 0),
                "duplicate_file_count": result.get("duplicate_file_count", 0),
                "unknown_file_count": result.get("unknown_file_count", 0),
                "safe_to_run_full_refresh": result.get("safe_to_run_full_refresh", False),
                "manifest_path": result.get("manifest_path", ""),
                "latest_manifest_path": result.get("latest_manifest_path", ""),
                "message": result.get("next_action", ""),
            }
        ],
    )


def _read_file_preview(file_path: Path) -> Dict[str, Any]:
    preview_rows_by_sheet = _load_preview_rows(file_path)
    patterns = load_report_patterns()
    synonyms = load_synonyms()

    best_result: Dict[str, Any] = {
        "detected_report_type": "unknown",
        "detected_period": _detect_period_from_filename(file_path.name),
        "warning": "",
        "headers": [],
        "sheet_name": "",
        "detected_columns": {},
    }
    best_score = -1.0

    for sheet_name, rows in preview_rows_by_sheet.items():
        if not rows:
            continue
        header_row_index, detected_columns, _ = detect_header_row(rows, synonyms, max_scan_rows=min(len(rows), 20))
        if header_row_index is None or header_row_index < 0 or header_row_index >= len(rows):
            headers = [normalize_text(cell) for cell in rows[0]]
        else:
            headers = [normalize_text(cell) for cell in rows[header_row_index]]
        report_type, scores = infer_report_type_scored(file_path.name, sheet_name, detected_columns, patterns)
        score = float(scores.get(report_type, 0.0)) if scores else 0.0
        category = _category_from_type(report_type, file_path.name, sheet_name, headers, detected_columns)
        if category == "unknown":
            score -= 0.25
        if score > best_score:
            best_score = score
            best_result = {
                "detected_report_type": category,
                "detected_period": _detect_period_from_filename(file_path.name),
                "warning": "",
                "headers": headers,
                "sheet_name": sheet_name,
                "detected_columns": detected_columns,
            }

    if best_score < 0:
        headers: List[str] = []
        if preview_rows_by_sheet:
            first_sheet_rows = next(iter(preview_rows_by_sheet.values()))
            if first_sheet_rows:
                headers = [normalize_text(cell) for cell in first_sheet_rows[0]]
        best_result = {
            "detected_report_type": _category_from_type("unknown", file_path.name, "", headers, {}),
            "detected_period": _detect_period_from_filename(file_path.name),
            "warning": "",
            "headers": headers,
            "sheet_name": "",
            "detected_columns": {},
        }

    if best_result["detected_report_type"] == "unknown":
        best_result["warning"] = _row_warning_text([best_result["warning"], "report type unknown from filename/header"])

    if best_result["detected_period"] == "Unknown":
        best_result["warning"] = _row_warning_text([best_result["warning"], "period unknown from filename"])

    return best_result


def _load_manifest_signature(path: Path) -> List[str]:
    if not path.exists():
        return []
    return _normalize_signature_rows_without_hash(_read_manifest_rows(path))


def check_flipkart_raw_input_safety() -> Dict[str, Any]:
    _ensure_output_dirs()

    manifest_rows: List[Dict[str, Any]] = []
    warnings: List[str] = []
    blockers: List[str] = []

    raw_exists = RAW_INPUT_DIR.exists()
    raw_files = list_input_files(RAW_INPUT_DIR) if raw_exists else []
    raw_file_count = len(raw_files)

    if not raw_exists:
        blockers.append(f"Raw input folder is missing: {RAW_INPUT_DIR}")
    elif raw_file_count == 0:
        blockers.append(f"Raw input folder is empty: {RAW_INPUT_DIR}")

    report_type_distribution: Counter[str] = Counter()
    period_distribution: Counter[str] = Counter()
    hash_to_files: defaultdict[str, List[str]] = defaultdict(list)
    report_type_to_dates: defaultdict[str, set[str]] = defaultdict(set)
    report_type_to_files: defaultdict[str, List[Tuple[Path, str, str]]] = defaultdict(list)
    duplicate_hash_count = 0
    unknown_file_count = 0

    for file_path in raw_files:
        file_size = file_path.stat().st_size
        modified_time = file_mtime_iso(file_path)
        modified_date = datetime.fromtimestamp(file_path.stat().st_mtime).date().isoformat()
        sha256_hash = _sha256_file(file_path)
        preview = _read_file_preview(file_path)
        detected_report_type = preview["detected_report_type"]
        detected_period = preview["detected_period"]
        row_warnings = [preview.get("warning", "")]

        report_type_distribution[detected_report_type] += 1
        period_distribution[detected_period] += 1
        hash_to_files[sha256_hash].append(file_path.name)
        report_type_to_dates[detected_report_type].add(modified_date)
        report_type_to_files[detected_report_type].append((file_path, modified_date, modified_time))

        if detected_report_type == "unknown":
            unknown_file_count += 1
        if detected_period == "Unknown":
            row_warnings.append("period unknown from filename")

        manifest_rows.append(
            {
                "file_name": file_path.name,
                "file_path": str(file_path),
                "file_size": file_size,
                "modified_time": modified_time,
                "sha256_hash": sha256_hash,
                "detected_report_type": detected_report_type,
                "detected_period": detected_period,
                "warning": _row_warning_text(row_warnings),
            }
        )

    duplicate_hash_groups = {hash_value: file_names for hash_value, file_names in hash_to_files.items() if len(file_names) > 1}
    duplicate_file_count = sum(len(file_names) - 1 for file_names in duplicate_hash_groups.values())
    if duplicate_hash_groups:
        blockers.append(f"Duplicate file hashes found for {len(duplicate_hash_groups)} hash group(s).")
        for hash_value, file_names in duplicate_hash_groups.items():
            warnings.append(f"Duplicate hash {hash_value[:12]}...: {', '.join(file_names)}")
            for row in manifest_rows:
                if row["sha256_hash"] == hash_value:
                    row["warning"] = _row_warning_text([row["warning"], "duplicate hash detected"])

    unknown_threshold = max(2, (raw_file_count + 3) // 4)
    if unknown_file_count > unknown_threshold:
        blockers.append(
            f"Too many unknown files: {unknown_file_count} of {raw_file_count} raw files (threshold {unknown_threshold})."
        )
    elif unknown_file_count > 0:
        warnings.append(f"{unknown_file_count} raw file(s) could not be categorized from filename/header.")

    clear_periods = sorted({period for period in period_distribution if period != "Unknown"})
    mixed_period_warning = len(clear_periods) > 1
    if mixed_period_warning:
        blockers.append(f"Multiple clear detected periods found: {', '.join(clear_periods)}")

    for report_type, dates in report_type_to_dates.items():
        if report_type == "unknown":
            continue
        if len(dates) > 1:
            sorted_dates = sorted(dates)
            blockers.append(
                f"Report type '{report_type}' appears across multiple modified dates: {', '.join(sorted_dates)}"
            )

    current_hashes = sorted(row["sha256_hash"] for row in manifest_rows)

    latest_manifest_previous = None
    if LATEST_MANIFEST_PATH.exists():
        try:
            latest_manifest_previous = json.loads(LATEST_MANIFEST_PATH.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            latest_manifest_previous = None

    previous_hashes = sorted(str(hash_value) for hash_value in latest_manifest_previous.get("manifest_hashes", [])) if isinstance(latest_manifest_previous, dict) else []
    same_manifest_as_previous_run = bool(previous_hashes and previous_hashes == current_hashes)
    if same_manifest_as_previous_run:
        blockers.append("Current raw manifest matches the latest saved raw manifest.")

    latest_run_manifest_path = _find_latest_run_manifest_path()
    latest_run_manifest_rows = _read_manifest_rows(latest_run_manifest_path) if latest_run_manifest_path else []
    latest_run_signature = _normalize_identity_rows(latest_run_manifest_rows) if latest_run_manifest_rows else []
    current_identity_rows = _normalize_identity_rows(manifest_rows)
    same_manifest_as_latest_run = bool(latest_run_signature and latest_run_signature == current_identity_rows)
    if same_manifest_as_latest_run:
        blockers.append("Current raw manifest matches the latest full run input manifest.")

    if raw_file_count == 0 and raw_exists:
        warnings.append("Raw input folder is present but contains no report files.")

    if raw_file_count > 0 and not manifest_rows:
        blockers.append("Unable to create a manifest for the current raw files.")

    safe_to_run_full_refresh = not blockers
    if safe_to_run_full_refresh and warnings:
        status = "PASS_WITH_WARNINGS"
    elif safe_to_run_full_refresh:
        status = "PASS"
    else:
        status = "BLOCKED"

    if blockers:
        warnings = blockers + warnings

    next_action = ""
    if status == "BLOCKED":
        if not raw_exists:
            next_action = f"Create {RAW_INPUT_DIR}, then place only current-cycle files there and rerun .\\check_flipkart_raw_input_safety.ps1."
        elif raw_file_count == 0:
            next_action = (
                f"Put only the current reporting cycle files into {RAW_INPUT_DIR}, archive old files to "
                f"data/input/marketplaces/flipkart/archive/YYYY-MM, and rerun .\\check_flipkart_raw_input_safety.ps1."
            )
        elif duplicate_hash_groups:
            next_action = (
                f"Remove duplicate copies or archive old duplicates from {RAW_INPUT_DIR}, then rerun "
                f".\\check_flipkart_raw_input_safety.ps1."
            )
        elif mixed_period_warning:
            next_action = (
                "Keep only one clear reporting period in the active raw folder, archive old cycle files to "
                "data/input/marketplaces/flipkart/archive/YYYY-MM, and rerun .\\check_flipkart_raw_input_safety.ps1."
            )
        elif same_manifest_as_previous_run or same_manifest_as_latest_run:
            next_action = (
                "Use a fresh raw reporting cycle before full refresh, or continue only with explicit "
                "--force-raw-refresh if you intentionally want to rerun the same manifest."
            )
        elif unknown_file_count > unknown_threshold:
            next_action = (
                "Rename or replace the unknown raw files so their report type is clear, then rerun "
                ".\\check_flipkart_raw_input_safety.ps1."
            )
        else:
            next_action = (
                "Archive mixed or older files into data/input/marketplaces/flipkart/archive/YYYY-MM, keep only the "
                "current cycle in raw, and rerun .\\check_flipkart_raw_input_safety.ps1."
            )
    else:
        next_action = ".\\run_flipkart_full_safe_refresh.ps1"

    manifest_payload = {
        "status": status,
        "generated_at": now_iso(),
        "raw_input_dir": str(RAW_INPUT_DIR),
        "raw_file_count": raw_file_count,
        "report_type_distribution": dict(sorted(report_type_distribution.items())),
        "period_distribution": dict(sorted(period_distribution.items())),
        "duplicate_file_count": duplicate_file_count,
        "unknown_file_count": unknown_file_count,
        "mixed_period_warning": mixed_period_warning,
        "same_manifest_as_previous_run": same_manifest_as_previous_run,
        "same_manifest_as_latest_run": same_manifest_as_latest_run,
        "safe_to_run_full_refresh": safe_to_run_full_refresh,
        "warnings": warnings,
        "next_action": next_action,
        "manifest_path": str(RAW_MANIFEST_PATH),
        "latest_manifest_path": str(LATEST_MANIFEST_PATH),
        "latest_run_manifest_path": str(latest_run_manifest_path) if latest_run_manifest_path else "",
        "manifest_hashes": current_hashes,
        "manifest_signature": current_identity_rows,
        "manifest_rows": manifest_rows,
    }

    write_csv(RAW_MANIFEST_PATH, MANIFEST_HEADERS, manifest_rows)
    _write_latest_manifest_json(manifest_payload)
    _append_log(manifest_payload)

    result = {
        "status": status,
        "raw_file_count": raw_file_count,
        "report_type_distribution": dict(sorted(report_type_distribution.items())),
        "period_distribution": dict(sorted(period_distribution.items())),
        "duplicate_file_count": duplicate_file_count,
        "unknown_file_count": unknown_file_count,
        "mixed_period_warning": mixed_period_warning,
        "same_manifest_as_previous_run": same_manifest_as_previous_run,
        "same_manifest_as_latest_run": same_manifest_as_latest_run,
        "safe_to_run_full_refresh": safe_to_run_full_refresh,
        "warnings": warnings,
        "next_action": next_action,
        "manifest_path": str(RAW_MANIFEST_PATH),
        "latest_manifest_path": str(LATEST_MANIFEST_PATH),
    }
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Check Flipkart raw input safety before a full monthly refresh.")
    parser.parse_args()

    try:
        result = check_flipkart_raw_input_safety()
        print(_json_text(result))
        if result["status"] == "BLOCKED":
            raise SystemExit(1)
    except SystemExit:
        raise
    except Exception as exc:
        error_payload = {
            "status": "ERROR",
            "error_type": exc.__class__.__name__,
            "message": str(exc),
            "manifest_path": str(RAW_MANIFEST_PATH),
            "latest_manifest_path": str(LATEST_MANIFEST_PATH),
        }
        _ensure_output_dirs()
        _write_latest_manifest_json(error_payload)
        print(_json_text(error_payload))
        raise SystemExit(1)


if __name__ == "__main__":
    main()

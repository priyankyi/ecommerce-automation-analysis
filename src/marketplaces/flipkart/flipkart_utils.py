from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

PROJECT_ROOT = Path(__file__).resolve().parents[3]
FLIPKART_ROOT = PROJECT_ROOT / "data" / "input" / "marketplaces" / "flipkart"
RAW_INPUT_DIR = FLIPKART_ROOT / "raw"
OUTPUT_DIR = PROJECT_ROOT / "data" / "output" / "marketplaces" / "flipkart"
LOG_DIR = PROJECT_ROOT / "data" / "logs" / "marketplaces" / "flipkart"
CONFIG_DIR = PROJECT_ROOT / "config"

ANALYSIS_JSON_PATH = OUTPUT_DIR / "report_analysis.json"
TARGET_FSN_PATH = OUTPUT_DIR / "flipkart_target_fsns.csv"
FSN_BRIDGE_PATH = OUTPUT_DIR / "flipkart_fsn_bridge.csv"
NORMALIZED_LISTINGS_PATH = OUTPUT_DIR / "normalized_listings.csv"
NORMALIZED_ORDERS_PATH = OUTPUT_DIR / "normalized_orders.csv"
NORMALIZED_RETURNS_PATH = OUTPUT_DIR / "normalized_returns.csv"
NORMALIZED_SETTLEMENTS_PATH = OUTPUT_DIR / "normalized_settlements.csv"
NORMALIZED_PNL_PATH = OUTPUT_DIR / "normalized_pnl.csv"
NORMALIZED_SALES_TAX_PATH = OUTPUT_DIR / "normalized_sales_tax.csv"
NORMALIZED_ADS_PATH = OUTPUT_DIR / "normalized_ads.csv"
SKU_ANALYSIS_PATH = OUTPUT_DIR / "flipkart_sku_analysis.csv"
NORMALIZATION_STATE_PATH = OUTPUT_DIR / "flipkart_normalization_state.json"
SKU_ANALYSIS_STATE_PATH = OUTPUT_DIR / "flipkart_sku_analysis_state.json"
PIPELINE_STATUS_PATH = OUTPUT_DIR / "flipkart_pipeline_status.json"

WORKBOOK_ROWS_CACHE: Dict[str, Dict[str, List[List[Any]]]] = {}

REPORT_ANALYSIS_LOG_PATH = LOG_DIR / "report_analysis_log.csv"
FSN_BRIDGE_LOG_PATH = LOG_DIR / "fsn_bridge_log.csv"
NORMALIZATION_LOG_PATH = LOG_DIR / "normalization_log.csv"
SKU_ANALYSIS_LOG_PATH = LOG_DIR / "flipkart_sku_analysis_log.csv"
PUSH_LOG_PATH = LOG_DIR / "push_to_sheet_log.csv"

TARGET_MASTER_FSN_FILENAME = "Master FSN File Fk (SPARKWORLD) (3).xlsx"
PANDAS_INSTALL_COMMAND = "python -m pip install pandas"
OPENPYXL_INSTALL_COMMAND = "python -m pip install openpyxl"
XLRD_INSTALL_COMMAND = "python -m pip install xlrd"

def ensure_directories() -> None:
    for path in [RAW_INPUT_DIR, OUTPUT_DIR, LOG_DIR, CONFIG_DIR]:
        path.mkdir(parents=True, exist_ok=True)


def load_json(path: Path) -> Dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Missing required file: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def load_run_state(path: Path) -> Dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Missing required file: {path}")
    return load_json(path)


def save_run_state(path: Path, payload: Dict[str, Any]) -> None:
    save_json(path, payload)


def write_csv(path: Path, fieldnames: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fieldnames))
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def append_csv_log(path: Path, fieldnames: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    file_exists = path.exists()
    with path.open("a", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fieldnames))
        if not file_exists:
            writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def now_iso_ms() -> str:
    return datetime.now().isoformat(timespec="milliseconds")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "null", "nan"}:
        return ""
    return text


def clean_fsn(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\ufeff", "")
    text = re.sub(r"[\u00A0\u200B\u200C\u200D]", "", text)
    text = re.sub(r"\s+", "", text).strip()
    if not text:
        return ""
    if text.lower() in {"none", "null", "nan"}:
        return ""
    return text.upper()


def normalize_key(value: Any) -> str:
    text = normalize_text(value).lower()
    text = text.replace("&", " and ")
    return re.sub(r"[^a-z0-9]+", "", text)


def is_blank(value: Any) -> bool:
    return normalize_text(value) == ""


def to_number(value: Any) -> Optional[float]:
    text = normalize_text(value)
    if not text:
        return None
    cleaned = text.replace(",", "")
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = f"-{cleaned[1:-1]}"
    cleaned = re.sub(r"[^0-9.\-]", "", cleaned)
    if cleaned in {"", "-", ".", "-."}:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def format_number(value: Any, decimals: int = 2) -> str:
    number = to_number(value)
    if number is None:
        return ""
    if number.is_integer():
        return str(int(number))
    return f"{number:.{decimals}f}".rstrip("0").rstrip(".")


def as_text(value: Any) -> str:
    return normalize_text(value)


def score_alias_match(header: str, alias: str) -> float:
    header_norm = normalize_key(header)
    alias_norm = normalize_key(alias)
    if not header_norm or not alias_norm:
        return 0.0
    if header_norm == alias_norm:
        return 1.2
    if alias_norm in header_norm or header_norm in alias_norm:
        return 0.92
    ratio = SequenceMatcher(None, header_norm, alias_norm).ratio()
    if ratio >= 0.8:
        return ratio
    header_tokens = set(re.findall(r"[a-z0-9]+", normalize_text(header).lower()))
    alias_tokens = set(re.findall(r"[a-z0-9]+", normalize_text(alias).lower()))
    if header_tokens and alias_tokens:
        overlap = len(header_tokens & alias_tokens) / len(alias_tokens)
        ratio = max(ratio, overlap)
    return ratio


def load_synonyms() -> Dict[str, List[str]]:
    return load_json(CONFIG_DIR / "flipkart_column_synonyms.json")


def load_report_patterns() -> Dict[str, Any]:
    return load_json(CONFIG_DIR / "flipkart_report_patterns.json")


def best_header_match(cell: Any, synonyms: Dict[str, List[str]]) -> Tuple[str, float]:
    text = normalize_text(cell)
    if not text:
        return "", 0.0
    best_key = ""
    best_score = 0.0
    for canonical, aliases in synonyms.items():
        for alias in aliases:
            score = score_alias_match(text, alias)
            if score > best_score:
                best_key = canonical
                best_score = score
    return best_key, best_score


def detect_header_row(rows: Sequence[Sequence[Any]], synonyms: Dict[str, List[str]], max_scan_rows: int = 25) -> Tuple[int, Dict[str, Dict[str, Any]], List[Dict[str, Any]]]:
    best_row_index = 0
    best_score = -1.0
    best_columns: Dict[str, Dict[str, Any]] = {}
    candidates: List[Dict[str, Any]] = []

    for row_index, row in enumerate(rows[:max_scan_rows]):
        matches: Dict[str, Dict[str, Any]] = {}
        total_score = 0.0
        for col_index, cell in enumerate(row):
            canonical, score = best_header_match(cell, synonyms)
            if canonical and score >= 0.82:
                current = matches.get(canonical)
                if current is None or score > current["score"]:
                    matches[canonical] = {
                        "header": normalize_text(cell),
                        "index": col_index,
                        "score": round(score, 4),
                    }
        total_score = sum(item["score"] for item in matches.values()) + len(matches) * 0.25
        candidates.append(
            {
                "row_index": row_index,
                "matched_columns": len(matches),
                "score": round(total_score, 4),
                "headers": [normalize_text(cell) for cell in row],
            }
        )
        if total_score > best_score:
            best_score = total_score
            best_row_index = row_index
            best_columns = matches

    if not best_columns and rows:
        fallback_matches: Dict[str, Dict[str, Any]] = {}
        for col_index, cell in enumerate(rows[0]):
            canonical, score = best_header_match(cell, synonyms)
            if canonical:
                fallback_matches[canonical] = {
                    "header": normalize_text(cell),
                    "index": col_index,
                    "score": round(score, 4),
                }
        best_columns = fallback_matches
        best_row_index = 0

    return best_row_index, best_columns, candidates


def likely_columns(headers: Sequence[Any], synonyms: Dict[str, List[str]], groups: Sequence[str]) -> Dict[str, Dict[str, Any]]:
    detected: Dict[str, Dict[str, Any]] = {}
    for canonical in groups:
        aliases = synonyms.get(canonical, [])
        best_score = 0.0
        best_header = ""
        best_index = -1
        for index, header in enumerate(headers):
            for alias in aliases:
                score = score_alias_match(header, alias)
                if score > best_score:
                    best_score = score
                    best_header = normalize_text(header)
                    best_index = index
        if best_score >= 0.82:
            detected[canonical] = {
                "header": best_header,
                "index": best_index,
                "score": round(best_score, 4),
            }
    return detected


def report_type_scores(file_name: str, sheet_name: str, detected_columns: Dict[str, Dict[str, Any]], patterns: Dict[str, Any]) -> Dict[str, float]:
    scores: Dict[str, float] = {}
    filename_norm = normalize_key(file_name)
    sheet_norm = normalize_key(sheet_name)
    for report_type, spec in patterns.get("report_types", {}).items():
        score = 0.0
        for keyword in spec.get("filename_keywords", []):
            keyword_norm = normalize_key(keyword)
            if keyword_norm and keyword_norm in filename_norm:
                score += 2.0
        for keyword in spec.get("sheet_keywords", []):
            keyword_norm = normalize_key(keyword)
            if keyword_norm and keyword_norm in sheet_norm:
                score += 1.0
        for canonical in spec.get("column_signals", []):
            if canonical in detected_columns:
                score += 1.5
        scores[report_type] = round(score, 4)
    return scores


def infer_report_type(file_name: str, sheet_name: str, detected_columns: Dict[str, Dict[str, Any]], patterns: Dict[str, Any]) -> Tuple[str, Dict[str, float]]:
    scores = report_type_scores(file_name, sheet_name, detected_columns, patterns)
    if not scores:
        return "unknown", scores
    best_type = max(scores, key=scores.get)
    if scores[best_type] <= 0:
        return "unknown", scores
    return best_type, scores


REPORT_TYPE_COLUMN_KEYS: Dict[str, Tuple[str, ...]] = {
    "listing": (
        "fsn",
        "sku_id",
        "product_title",
        "category",
        "listing_status",
        "inactive_reason",
        "mrp",
        "selling_price",
        "stock",
        "bank_settlement",
        "listing_quality",
        "package_length",
        "package_breadth",
        "package_height",
        "package_weight",
        "hsn",
        "tax_code",
    ),
    "orders": (
        "fsn",
        "order_item_id",
        "order_id",
        "sku_id",
        "product_title",
        "quantity",
        "selling_price",
        "order_status",
        "dispatch_date",
        "delivery_date",
        "cancellation_status",
        "order_date",
    ),
    "returns": (
        "fsn",
        "order_item_id",
        "order_id",
        "sku_id",
        "return_id",
        "return_date",
        "return_type",
        "return_reason",
        "return_status",
        "reverse_shipment_status",
    ),
    "settlements": (
        "fsn",
        "order_item_id",
        "order_id",
        "sku_id",
        "settlement_id",
        "settlement_date",
        "gross_amount",
        "commission",
        "fixed_fee",
        "collection_fee",
        "shipping_fee",
        "reverse_shipping_fee",
        "gst_on_fees",
        "tcs",
        "tds",
        "refund",
        "protection_fund",
        "adjustments",
        "net_settlement",
    ),
    "pnl": (
        "fsn",
        "sku_id",
        "order_item_id",
        "order_id",
        "flipkart_net_earnings",
        "flipkart_margin",
        "flipkart_expenses",
        "amount_settled",
        "amount_pending",
        "revenue",
        "expenses",
        "net_earnings",
        "margin",
    ),
    "sales_tax": (
        "fsn",
        "order_item_id",
        "order_id",
        "sku_id",
        "invoice_id",
        "invoice_date",
        "hsn",
        "taxable_value",
        "igst",
        "cgst",
        "sgst",
        "tcs",
        "tds",
        "event_type",
    ),
    "ads": (
        "fsn",
        "sku_id",
        "campaign_id",
        "campaign_name",
        "adgroup_id",
        "adgroup_name",
        "views",
        "clicks",
        "direct_units_sold",
        "indirect_units_sold",
        "total_revenue",
        "roi",
        "estimated_ad_spend",
        "roas",
        "acos",
        "product_name",
    ),
}

IDENTIFIER_COLUMN_KEYS = ("fsn", "order_item_id", "order_id", "sku_id", "settlement_id", "invoice_id")
MONEY_COLUMN_KEYS = (
    "mrp",
    "selling_price",
    "bank_settlement",
    "gross_amount",
    "commission",
    "fixed_fee",
    "collection_fee",
    "shipping_fee",
    "reverse_shipping_fee",
    "gst_on_fees",
    "tcs",
    "tds",
    "refund",
    "protection_fund",
    "adjustments",
    "net_settlement",
    "flipkart_net_earnings",
    "flipkart_margin",
    "flipkart_expenses",
    "amount_settled",
    "amount_pending",
    "revenue",
    "expenses",
    "net_earnings",
    "margin",
    "taxable_value",
    "igst",
    "cgst",
    "sgst",
    "total_revenue",
    "estimated_ad_spend",
)
DATE_COLUMN_KEYS = ("order_date", "dispatch_date", "delivery_date", "return_date", "settlement_date", "invoice_date")
STATUS_COLUMN_KEYS = ("listing_status", "order_status", "return_status", "cancellation_status", "reverse_shipment_status", "event_type", "inactive_reason")
ALL_USEFUL_COLUMN_KEYS = tuple(sorted(set().union(*[set(keys) for keys in REPORT_TYPE_COLUMN_KEYS.values()])))


def count_non_empty_rows(rows: Sequence[Sequence[Any]]) -> int:
    return sum(1 for row in rows if any(not is_blank(cell) for cell in row))


def count_detected_columns(detected_columns: Dict[str, Dict[str, Any]], keys: Sequence[str]) -> int:
    return sum(1 for key in keys if key in detected_columns)


def sheet_exclusion_reason(sheet_name: str, detected_columns: Dict[str, Dict[str, Any]]) -> str:
    sheet_norm = normalize_key(sheet_name)
    if any(token in sheet_norm for token in ("reporthelp", "help", "instructions", "readme")):
        return "sheet name indicates help/instructions/readme"
    if "summary" in sheet_norm and count_detected_columns(detected_columns, ALL_USEFUL_COLUMN_KEYS) == 0:
        return "summary sheet without transactional columns"
    return ""


def sheet_selection_metrics(
    file_name: str,
    sheet_name: str,
    report_type: str,
    rows: Sequence[Sequence[Any]],
    header_row_index: int,
    detected_columns: Dict[str, Dict[str, Any]],
    patterns: Dict[str, Any],
) -> Dict[str, Any]:
    data_rows = rows[header_row_index + 1 :] if rows else []
    non_empty_rows = count_non_empty_rows(data_rows)
    useful_keys = REPORT_TYPE_COLUMN_KEYS.get(report_type, ())
    useful_column_count = count_detected_columns(detected_columns, useful_keys)
    identifier_hits = count_detected_columns(detected_columns, IDENTIFIER_COLUMN_KEYS)
    money_hits = count_detected_columns(detected_columns, MONEY_COLUMN_KEYS)
    date_hits = count_detected_columns(detected_columns, DATE_COLUMN_KEYS)
    status_hits = count_detected_columns(detected_columns, STATUS_COLUMN_KEYS)
    exclusion_reason = sheet_exclusion_reason(sheet_name, detected_columns)
    type_scores = report_type_scores(file_name, sheet_name, detected_columns, patterns)
    report_score = type_scores.get(report_type, 0.0)

    selection_score = (
        (non_empty_rows * 1.0)
        + (useful_column_count * 8.0)
        + (identifier_hits * 12.0)
        + (money_hits * 4.0)
        + (date_hits * 2.0)
        + (status_hits * 2.0)
        + report_score
    )
    eligible = not exclusion_reason and useful_column_count > 0

    return {
        "file_name": file_name,
        "sheet_name": sheet_name,
        "report_type": report_type,
        "header_row_index": header_row_index,
        "raw_row_count": len(data_rows),
        "non_empty_row_count": non_empty_rows,
        "useful_column_count": useful_column_count,
        "identifier_hits": identifier_hits,
        "money_hits": money_hits,
        "date_hits": date_hits,
        "status_hits": status_hits,
        "exclusion_reason": exclusion_reason,
        "report_type_score": report_score,
        "selection_score": round(selection_score, 4),
        "eligible": eligible,
    }


def score_sheet_candidate(
    file_name: str,
    sheet_name: str,
    rows: Sequence[Sequence[Any]],
    report_type: str,
    synonyms: Dict[str, List[str]],
    patterns: Dict[str, Any],
    max_scan_rows: int = 50,
) -> Dict[str, Any]:
    header_row_index, detected_columns, candidates = detect_header_row(rows, synonyms, max_scan_rows=max_scan_rows)
    metrics = sheet_selection_metrics(file_name, sheet_name, report_type, rows, header_row_index, detected_columns, patterns)
    return {
        **metrics,
        "header_row": ["" if cell is None else str(cell).strip() for cell in rows[header_row_index]] if rows else [],
        "detected_columns": detected_columns,
        "header_candidates": candidates,
    }


def select_best_sheet_candidate(candidates: Sequence[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    eligible = [candidate for candidate in candidates if candidate.get("eligible")]
    if not eligible:
        return None
    return max(
        eligible,
        key=lambda candidate: (
            float(candidate.get("selection_score", 0.0)),
            int(candidate.get("useful_column_count", 0)),
            int(candidate.get("non_empty_row_count", 0)),
            int(candidate.get("identifier_hits", 0)),
            int(candidate.get("money_hits", 0)),
            -int(candidate.get("header_row_index", 0)),
        ),
    )


def select_best_sheet_for_report(
    file_name: str,
    workbook_rows: Dict[str, List[List[Any]]],
    report_type: str,
    synonyms: Dict[str, List[str]],
    patterns: Dict[str, Any],
    max_scan_rows: int = 50,
) -> Dict[str, Any]:
    candidates: List[Dict[str, Any]] = []
    for sheet_name, rows in workbook_rows.items():
        if not rows:
            candidates.append(
                {
                    "file_name": file_name,
                    "sheet_name": sheet_name,
                    "report_type": report_type,
                    "header_row_index": 0,
                    "raw_row_count": 0,
                    "non_empty_row_count": 0,
                    "useful_column_count": 0,
                    "identifier_hits": 0,
                    "money_hits": 0,
                    "date_hits": 0,
                    "status_hits": 0,
                    "exclusion_reason": "empty sheet",
                    "report_type_score": 0.0,
                    "selection_score": 0.0,
                    "eligible": False,
                    "header_row": [],
                    "detected_columns": {},
                    "header_candidates": [],
                }
            )
            continue
        candidates.append(score_sheet_candidate(file_name, sheet_name, rows, report_type, synonyms, patterns, max_scan_rows=max_scan_rows))

    selected = select_best_sheet_candidate(candidates)
    rejected = [candidate for candidate in candidates if candidate is not selected]
    return {
        "file_name": file_name,
        "report_type": report_type,
        "selected_sheet": selected,
        "candidates": candidates,
        "rejected_sheets": rejected,
    }


def select_best_sheet_across_files(
    file_paths: Sequence[Path],
    report_type: str,
    synonyms: Dict[str, List[str]],
    patterns: Dict[str, Any],
    max_scan_rows: int = 50,
) -> Dict[str, Any]:
    best_selection: Optional[Dict[str, Any]] = None
    all_candidates: List[Dict[str, Any]] = []
    all_rejected: List[Dict[str, Any]] = []
    hinted_selections: List[Dict[str, Any]] = []
    fallback_selections: List[Dict[str, Any]] = []

    for file_path in file_paths:
        workbook_rows = read_workbook_rows(file_path)
        selection = select_best_sheet_for_report(file_path.name, workbook_rows, report_type, synonyms, patterns, max_scan_rows=max_scan_rows)
        file_hint_score = report_type_scores(file_path.name, "", {}, patterns).get(report_type, 0.0)
        selected = selection.get("selected_sheet")
        if selected:
            selected = dict(selected)
            selected["file_hint_score"] = file_hint_score
            selected["combined_score"] = round((file_hint_score * 100.0) + float(selected.get("selection_score", 0.0)), 4)
            selected["file_path"] = str(file_path)
            selection["selected_sheet"] = selected
            if file_hint_score > 0:
                hinted_selections.append(selection)
            else:
                fallback_selections.append(selection)
        for candidate in selection.get("candidates", []):
            candidate["file_hint_score"] = file_hint_score
            candidate["combined_score"] = round((file_hint_score * 100.0) + float(candidate.get("selection_score", 0.0)), 4)
            candidate["file_path"] = str(file_path)
            all_candidates.append(candidate)
        for rejected in selection.get("rejected_sheets", []):
            rejected["file_hint_score"] = file_hint_score
            rejected["combined_score"] = round((file_hint_score * 100.0) + float(rejected.get("selection_score", 0.0)), 4)
            rejected["file_path"] = str(file_path)
            all_rejected.append(rejected)

    ranked_selections = hinted_selections
    if not ranked_selections:
        return {
            "report_type": report_type,
            "selected_sheet": None,
            "candidates": all_candidates,
            "rejected_sheets": all_rejected,
        }
    for selection in ranked_selections:
        selected = selection.get("selected_sheet")
        if not selected:
            continue
        if best_selection is None or float(selected["combined_score"]) > float(best_selection["selected_sheet"]["combined_score"]):
            best_selection = selection

    if best_selection is None:
        return {
            "report_type": report_type,
            "selected_sheet": None,
            "candidates": all_candidates,
            "rejected_sheets": all_rejected,
        }

    best_selected = best_selection.get("selected_sheet")
    if best_selected:
        best_selection = dict(best_selection)
        best_selection["selected_sheet"] = best_selected
    best_selection["candidates"] = all_candidates
    best_selection["rejected_sheets"] = all_rejected
    return best_selection


def get_cell(row: Sequence[Any], index: int) -> Any:
    if index < 0 or index >= len(row):
        return ""
    return row[index]


def collect_values(row: Sequence[Any], detected_columns: Dict[str, Dict[str, Any]], keys: Sequence[str]) -> Dict[str, str]:
    values: Dict[str, str] = {}
    for key in keys:
        column = detected_columns.get(key)
        if column is None:
            values[key] = ""
        else:
            values[key] = as_text(get_cell(row, int(column["index"])))
    return values


def read_csv_rows(path: Path) -> List[List[Any]]:
    try:
        import pandas as pd
    except ImportError as exc:  # pragma: no cover - dependency failure
        raise RuntimeError(f"Missing dependency `pandas`. Install it with: {PANDAS_INSTALL_COMMAND}") from exc
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        raw_lines = [line.rstrip("\n") for line in handle]

    if not raw_lines:
        return []

    header_index = 0
    header_score = -1
    for index, line in enumerate(raw_lines[:200]):
        comma_count = line.count(",")
        if comma_count > header_score:
            header_score = comma_count
            header_index = index

    frame = pd.read_csv(
        path,
        dtype=str,
        keep_default_na=False,
        encoding="utf-8-sig",
        skiprows=header_index,
        header=0,
        engine="python",
        on_bad_lines="skip",
    )
    return [list(frame.columns)] + frame.fillna("").values.tolist()


def read_workbook_rows(path: Path) -> Dict[str, List[List[Any]]]:
    if not path.exists():
        raise FileNotFoundError(f"Missing required file: {path}")
    cache_key = str(path.resolve())
    cached = WORKBOOK_ROWS_CACHE.get(cache_key)
    if cached is not None:
        return cached
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = {path.name: read_csv_rows(path)}
        WORKBOOK_ROWS_CACHE[cache_key] = rows
        return rows
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency failure
            raise RuntimeError(f"Missing dependency `openpyxl`. Install it with: {OPENPYXL_INSTALL_COMMAND}") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_rows: Dict[str, List[List[Any]]] = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            sheet_rows[sheet_name] = rows
        WORKBOOK_ROWS_CACHE[cache_key] = sheet_rows
        return sheet_rows
    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover - dependency failure
            raise RuntimeError(f"Missing dependency `xlrd`. Install it with: {XLRD_INSTALL_COMMAND}") from exc

        workbook = xlrd.open_workbook(str(path))
        sheet_rows = {}
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            rows: List[List[Any]] = []
            for row_index in range(sheet.nrows):
                values = sheet.row_values(row_index)
                rows.append(values)
            sheet_rows[sheet_name] = rows
        WORKBOOK_ROWS_CACHE[cache_key] = sheet_rows
        return sheet_rows
    return {}


def list_input_files(input_dir: Path) -> List[Path]:
    if not input_dir.exists():
        return []
    files: List[Path] = []
    for path in sorted(input_dir.iterdir()):
        if path.name.startswith("~$"):
            continue
        if path.suffix.lower() in {".csv", ".xls", ".xlsx", ".xlsm"}:
            files.append(path)
    return files


def build_file_sheet_id(file_name: str, sheet_name: str) -> str:
    return f"{file_name}::{sheet_name}"


def choose_confidence(levels: Sequence[str]) -> str:
    priority = {"HIGH": 3, "MEDIUM": 2, "LOW": 1, "UNKNOWN": 0}
    best = "UNKNOWN"
    for level in levels:
        level_norm = normalize_text(level).upper() or "UNKNOWN"
        if priority.get(level_norm, 0) > priority.get(best, 0):
            best = level_norm
    return best


def highest_priority_fsn(candidates: Sequence[str], priority_map: Dict[str, Any]) -> str:
    best_fsn = ""
    best_rank = float("inf")
    for fsn in candidates:
        if not fsn:
            continue
        priority_value = priority_map.get(fsn, {}).get("priority")
        try:
            rank = float(priority_value)
        except Exception:
            rank = float("inf")
        if rank < best_rank:
            best_rank = rank
            best_fsn = fsn
    if best_fsn:
        return best_fsn
    return sorted([fsn for fsn in candidates if fsn])[:1][0] if [fsn for fsn in candidates if fsn] else ""


def parse_float(value: Any) -> float:
    number = to_number(value)
    return number if number is not None else 0.0


def parse_int(value: Any) -> int:
    number = to_number(value)
    if number is None:
        return 0
    return int(number)


def format_decimal(value: Any, decimals: int = 2) -> str:
    number = to_number(value)
    if number is None:
        return ""
    return f"{number:.{decimals}f}"


def dedupe_dict_rows(rows: Sequence[Dict[str, Any]], key_field: str) -> List[Dict[str, Any]]:
    seen = set()
    result = []
    for row in rows:
        key = normalize_text(row.get(key_field))
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(row)
    return result


def merge_non_blank(existing: Dict[str, Any], candidate: Dict[str, Any], fields: Sequence[str]) -> None:
    for field in fields:
        if is_blank(existing.get(field)) and not is_blank(candidate.get(field)):
            existing[field] = candidate[field]


def record_count(rows: Sequence[Any]) -> int:
    return len(rows)


def csv_data_row_count(path: Path) -> int:
    if not path.exists():
        return 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return max(0, sum(1 for _ in csv.reader(handle)) - 1)


def file_mtime_iso(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds")


def path_mtime(path: Path) -> float:
    return path.stat().st_mtime


def build_status_payload(status: str, **extra: Any) -> Dict[str, Any]:
    payload = {"status": status, "generated_at": now_iso_ms()}
    extra.pop("status", None)
    payload.update(extra)
    return payload

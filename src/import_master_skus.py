from __future__ import annotations

import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.auth_google import build_services

SPREADSHEET_META_PATH = PROJECT_ROOT / "data" / "output" / "master_sku_sheet.json"
SOURCE_TEMPLATE_PATH = PROJECT_ROOT / "data" / "input" / "sku_import_template.csv"
SOURCE_DIR = PROJECT_ROOT / "data" / "input" / "master_skus"
LOG_PATH = PROJECT_ROOT / "data" / "logs" / "master_sku_import_log.csv"
MASTER_TAB = "MASTER_SKU"

REQUIRED_FIELDS = [
    "SKU_ID",
    "Product_Title",
    "Category",
    "Cost_Price",
    "MRP",
    "Selling_Price",
    "GST_Rate",
    "HSN_Code",
    "Length_cm",
    "Width_cm",
    "Height_cm",
    "Dead_Weight_kg",
    "Available_Stock",
    "Reorder_Level",
    "Product_Status",
]

OPTIONAL_TIMESTAMP_FIELDS = ["Created_Date", "Last_Updated"]
FORMULA_HEADERS = ["Volumetric_Weight_kg", "Final_Chargeable_Weight"]


def load_json(path: Path) -> Dict[str, object]:
    if not path.exists():
        raise FileNotFoundError(f"Missing required file: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def column_index_to_a1(index: int) -> str:
    if index < 1:
        raise ValueError("Column index must be 1 or greater")
    result = []
    while index:
        index, remainder = divmod(index - 1, 26)
        result.append(chr(65 + remainder))
    return "".join(reversed(result))


def zero_based_column_index_to_a1(index: int) -> str:
    return column_index_to_a1(index + 1)


def retry(func, attempts: int = 4):
    delay = 1.0
    for attempt in range(1, attempts + 1):
        try:
            return func()
        except Exception:
            if attempt == attempts:
                raise
            import time

            time.sleep(delay)
            delay *= 2


def get_sheet_values(sheets_service, spreadsheet_id: str, range_name: str) -> List[List[str]]:
    return retry(
        lambda: sheets_service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=range_name)
        .execute()
        .get("values", [])
    )


def find_header_index(headers: Sequence[str], header_name: str) -> int | None:
    for index, header in enumerate(headers):
        if header == header_name:
            return index
    return None


def read_csv_file(path: Path) -> Tuple[List[str], List[List[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def read_xlsx_file(path: Path) -> Tuple[List[str], List[List[str]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required for XLSX imports. Install dependencies with pip install -r requirements.txt"
        ) from exc

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = ["" if cell is None else str(cell).strip() for cell in rows[0]]
    data_rows: List[List[str]] = []
    for row in rows[1:]:
        data_rows.append(["" if cell is None else str(cell).strip() for cell in row])
    return headers, data_rows


def read_input_file(path: Path) -> Tuple[List[str], List[List[str]]]:
    if path.suffix.lower() == ".csv":
        return read_csv_file(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_xlsx_file(path)
    return [], []


def source_files() -> List[Path]:
    files: List[Path] = []
    if SOURCE_TEMPLATE_PATH.exists():
        files.append(SOURCE_TEMPLATE_PATH)
    if SOURCE_DIR.exists():
        files.extend(sorted(path for path in SOURCE_DIR.iterdir() if path.suffix.lower() in {".csv", ".xlsx", ".xlsm"}))
    return files


def clean_text(value: object) -> str:
    text = "" if value is None else str(value).strip()
    if text.lower() in {"none", "null"}:
        return ""
    return text


def build_source_row(headers: List[str], row: List[str]) -> Dict[str, str]:
    return {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}


def validate_required_fields(source_row: Dict[str, str]) -> List[str]:
    errors: List[str] = []
    for field in REQUIRED_FIELDS:
        if not clean_text(source_row.get(field)):
            errors.append(f"Missing required field: {field}")
    return errors


def get_existing_skus(sheets_service, spreadsheet_id: str) -> Tuple[List[str], List[str], List[List[str]]]:
    rows = get_sheet_values(sheets_service, spreadsheet_id, f"{MASTER_TAB}!A1:ZZ")
    if not rows:
        raise ValueError("MASTER_SKU sheet is empty or missing headers.")

    headers = rows[0]
    sku_index = find_header_index(headers, "SKU_ID")
    if sku_index is None:
        raise ValueError("SKU_ID header was not found in MASTER_SKU.")

    existing: List[str] = []
    for row in rows[1:]:
        sku_id = row[sku_index].strip() if sku_index < len(row) else ""
        if sku_id:
            existing.append(sku_id)
    return headers, existing, rows


def build_formula(header_map: Dict[str, int], header: str, row_number: int) -> str | None:
    if header == "Volumetric_Weight_kg":
        if not all(name in header_map for name in ["Length_cm", "Width_cm", "Height_cm"]):
            return None
        length = column_index_to_a1(header_map["Length_cm"] + 1)
        width = column_index_to_a1(header_map["Width_cm"] + 1)
        height = column_index_to_a1(header_map["Height_cm"] + 1)
        return f"=ROUND(({length}{row_number}*{width}{row_number}*{height}{row_number})/5000, 2)"
    if header == "Final_Chargeable_Weight":
        if not all(name in header_map for name in ["Dead_Weight_kg", "Volumetric_Weight_kg"]):
            return None
        dead = column_index_to_a1(header_map["Dead_Weight_kg"] + 1)
        vol = column_index_to_a1(header_map["Volumetric_Weight_kg"] + 1)
        return f"=MAX({dead}{row_number}, {vol}{row_number})"
    return None


def build_row(
    source_row: Dict[str, str],
    headers: List[str],
    row_number: int,
    timestamp: str,
) -> List[str]:
    header_map = {header: index for index, header in enumerate(headers)}
    row = [""] * len(headers)

    for header, value in source_row.items():
        if header in header_map:
            row[header_map[header]] = value

    for header in OPTIONAL_TIMESTAMP_FIELDS:
        if header in header_map and not clean_text(row[header_map[header]]):
            row[header_map[header]] = timestamp

    for header in FORMULA_HEADERS:
        if header in header_map:
            formula = build_formula(header_map, header, row_number)
            if formula:
                row[header_map[header]] = formula

    return row


def append_log_rows(log_rows: List[Dict[str, object]]) -> None:
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    file_exists = LOG_PATH.exists()
    with LOG_PATH.open("a", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["timestamp", "source_file", "row_number", "sku_id", "status", "message"],
        )
        if not file_exists:
            writer.writeheader()
        writer.writerows(log_rows)


def import_master_skus() -> Dict[str, object]:
    meta = load_json(SPREADSHEET_META_PATH)
    spreadsheet_id = meta["spreadsheet_id"]
    sheets_service, _, _ = build_services()

    headers, existing_skus, existing_rows = get_existing_skus(sheets_service, spreadsheet_id)
    header_map = {header: index for index, header in enumerate(headers)}
    missing_required_in_sheet = [field for field in REQUIRED_FIELDS if field not in header_map]
    if missing_required_in_sheet:
        raise ValueError(f"MASTER_SKU is missing required sheet columns: {', '.join(missing_required_in_sheet)}")

    files = source_files()
    imported_rows: List[List[str]] = []
    log_rows: List[Dict[str, object]] = []
    validation_errors = 0
    duplicate_skipped = 0
    rows_read = 0
    seen_in_batch: set[str] = set(existing_skus)
    timestamp = datetime.now().isoformat(timespec="seconds")
    next_row_number = len(existing_rows) + 1

    for path in files:
        source_headers, data_rows = read_input_file(path)
        print(f"Found file: {path.name} | rows read: {len(data_rows)}")
        if not source_headers:
            continue

        for index, raw_row in enumerate(data_rows, start=2):
            rows_read += 1
            source_row = build_source_row(source_headers, raw_row)
            sku_id = clean_text(source_row.get("SKU_ID"))
            row_errors = validate_required_fields(source_row)

            if not sku_id:
                row_errors.append("Missing required field: SKU_ID")

            if sku_id in seen_in_batch:
                duplicate_skipped += 1
                log_rows.append(
                    {
                        "timestamp": timestamp,
                        "source_file": path.name,
                        "row_number": index,
                        "sku_id": sku_id,
                        "status": "skipped",
                        "message": "Duplicate SKU_ID",
                    }
                )
                continue

            if row_errors:
                validation_errors += 1
                log_rows.append(
                    {
                        "timestamp": timestamp,
                        "source_file": path.name,
                        "row_number": index,
                        "sku_id": sku_id,
                        "status": "invalid",
                        "message": "; ".join(row_errors),
                    }
                )
                continue

            imported_rows.append(build_row(source_row, headers, next_row_number, timestamp))
            seen_in_batch.add(sku_id)
            log_rows.append(
                {
                    "timestamp": timestamp,
                    "source_file": path.name,
                    "row_number": index,
                    "sku_id": sku_id,
                    "status": "created",
                    "message": "Imported MASTER_SKU row",
                }
            )
            next_row_number += 1

    if imported_rows:
        sheets_service.spreadsheets().values().append(
            spreadsheetId=spreadsheet_id,
            range=f"{MASTER_TAB}!A:ZZ",
            valueInputOption="USER_ENTERED",
            insertDataOption="INSERT_ROWS",
            body={"values": imported_rows},
        ).execute()

    append_log_rows(log_rows)

    print("Import summary:")
    print(f"  Files found: {len(files)}")
    print(f"  Rows read: {rows_read}")
    print(f"  Rows imported: {len(imported_rows)}")
    print(f"  Duplicates skipped: {duplicate_skipped}")
    print(f"  Validation errors: {validation_errors}")

    return {
        "spreadsheet_id": spreadsheet_id,
        "files_found": len(files),
        "rows_read": rows_read,
        "rows_imported": len(imported_rows),
        "duplicates_skipped": duplicate_skipped,
        "validation_errors": validation_errors,
        "log_path": str(LOG_PATH),
    }


def main() -> None:
    result = import_master_skus()
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()

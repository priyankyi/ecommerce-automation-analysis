from __future__ import annotations

import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.auth_google import build_services

SPREADSHEET_META_PATH = PROJECT_ROOT / "data" / "output" / "master_sku_sheet.json"
ADS_INPUT_DIR = PROJECT_ROOT / "data" / "input" / "ads"
MAPPING_PATH = PROJECT_ROOT / "config" / "ads_column_mapping.json"
LOG_PATH = PROJECT_ROOT / "data" / "logs" / "ads_import_log.csv"

RAW_TAB = "ADS_RAW"
MASTER_TAB = "ADS_MASTER"

RAW_HEADERS = [
    "Source_File",
    "Marketplace",
    "Import_Timestamp",
    "Raw_Campaign_ID",
    "Raw_Campaign_Name",
    "Raw_SKU",
    "Raw_Product_Title",
    "Raw_Date",
    "Raw_Impressions",
    "Raw_Clicks",
    "Raw_Ad_Spend",
    "Raw_Ad_Sales",
    "Raw_Orders",
    "Raw_Data_JSON",
]

MASTER_HEADERS = [
    "Campaign_ID",
    "Campaign_Name",
    "Marketplace",
    "SKU_ID",
    "Product_Title",
    "Date",
    "Impressions",
    "Clicks",
    "Ad_Spend",
    "Ad_Sales",
    "Orders",
    "CTR",
    "CPC",
    "ROAS",
    "ACOS",
    "Import_Source_File",
    "Last_Updated",
]


def load_json(path: Path) -> Dict[str, object]:
    if not path.exists():
        raise FileNotFoundError(f"Missing required file: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def column_index_to_a1(index: int) -> str:
    if index < 1:
        raise ValueError("Column index must be 1 or greater")
    result = []
    while index:
        index, remainder = divmod(index - 1, 26)
        result.append(chr(65 + remainder))
    return "".join(reversed(result))


def zero_based_column_index_to_a1(index: int) -> str:
    return column_index_to_a1(index + 1)


def get_metadata(sheets_service, spreadsheet_id: str) -> Dict[str, object]:
    return (
        sheets_service.spreadsheets()
        .get(spreadsheetId=spreadsheet_id, fields="sheets(properties(sheetId,title))")
        .execute()
    )


def find_tab(metadata: Dict[str, object], tab_name: str) -> Dict[str, object] | None:
    for sheet in metadata.get("sheets", []):
        properties = sheet.get("properties", {})
        if properties.get("title") == tab_name:
            return properties
    return None


def ensure_tab(sheets_service, spreadsheet_id: str, tab_name: str) -> int:
    metadata = get_metadata(sheets_service, spreadsheet_id)
    existing = find_tab(metadata, tab_name)
    if existing and existing.get("sheetId") is not None:
        return existing["sheetId"]

    response = (
        sheets_service.spreadsheets()
        .batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": [{"addSheet": {"properties": {"title": tab_name}}}]},
        )
        .execute()
    )
    return response["replies"][0]["addSheet"]["properties"]["sheetId"]


def get_sheet_values(sheets_service, spreadsheet_id: str, range_name: str) -> List[List[str]]:
    response = (
        sheets_service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=range_name)
        .execute()
    )
    return response.get("values", [])


def find_header_index(headers: Sequence[str], header_name: str) -> int | None:
    for index, header in enumerate(headers):
        if header == header_name:
            return index
    return None


def ensure_headers(sheets_service, spreadsheet_id: str, tab_name: str, headers: List[str]) -> bool:
    current = get_sheet_values(sheets_service, spreadsheet_id, f"{tab_name}!A1:Z1")
    current_headers = current[0] if current else []
    if current_headers[: len(headers)] != headers:
        sheets_service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"{tab_name}!A1:{zero_based_column_index_to_a1(len(headers) - 1)}1",
            valueInputOption="RAW",
            body={"values": [headers]},
        ).execute()
        return True
    return False


def load_mapping() -> Dict[str, object]:
    return load_json(MAPPING_PATH)


def resolve_column(headers: Sequence[str], aliases: Iterable[str]) -> int | None:
    for alias in aliases:
        index = find_header_index(headers, alias)
        if index is not None:
            return index
    return None


def clean_text(value: object) -> str:
    text = "" if value is None else str(value).strip()
    if text.lower() in {"none", "null"}:
        return ""
    return text


def to_float(value: object) -> float:
    text = clean_text(value)
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def to_int_like(value: object) -> str:
    number = to_float(value)
    if number.is_integer():
        return str(int(number))
    return f"{number}"


def format_ratio(numerator: float, denominator: float) -> str:
    if denominator == 0:
        return "0"
    return f"{numerator / denominator:.4f}"


def format_currency_ratio(numerator: float, denominator: float) -> str:
    if denominator == 0:
        return "0"
    return f"{numerator / denominator:.4f}"


def read_csv_file(path: Path) -> Tuple[List[str], List[List[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def read_xlsx_file(path: Path) -> Tuple[List[str], List[List[str]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for XLSX imports. Install dependencies with pip install -r requirements.txt") from exc

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = ["" if cell is None else str(cell).strip() for cell in rows[0]]
    data_rows: List[List[str]] = []
    for row in rows[1:]:
        data_rows.append(["" if cell is None else str(cell).strip() for cell in row])
    return headers, data_rows


def read_input_file(path: Path) -> Tuple[List[str], List[List[str]]]:
    if path.suffix.lower() == ".csv":
        return read_csv_file(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_xlsx_file(path)
    return [], []


def normalize_row(
    source_file: str,
    raw_headers: Sequence[str],
    raw_row: Sequence[str],
    mapping: Dict[str, object],
) -> Tuple[List[str], List[str], Dict[str, str]]:
    column_map = mapping["columns"]
    default_marketplace = mapping.get("default_marketplace", "Unknown")
    raw_lookup = {header: (raw_row[i] if i < len(raw_row) else "") for i, header in enumerate(raw_headers)}

    def get_value(key: str) -> str:
        aliases = column_map.get(key, [])
        index = resolve_column(raw_headers, aliases)
        if index is not None and index < len(raw_row):
            return clean_text(raw_row[index])
        return ""

    campaign_id = get_value("campaign_id")
    campaign_name = get_value("campaign_name")
    marketplace = get_value("marketplace") or default_marketplace
    sku_id = get_value("sku_id")
    product_title = get_value("product_title")
    date = get_value("date")
    impressions = to_float(get_value("impressions"))
    clicks = to_float(get_value("clicks"))
    ad_spend = to_float(get_value("ad_spend"))
    ad_sales = to_float(get_value("ad_sales"))
    orders = to_float(get_value("orders"))

    ctr = format_ratio(clicks, impressions)
    cpc = format_currency_ratio(ad_spend, clicks)
    roas = format_ratio(ad_sales, ad_spend)
    acos = format_ratio(ad_spend, ad_sales)

    import_timestamp = datetime.now().isoformat(timespec="seconds")
    raw_json = json.dumps(raw_lookup, ensure_ascii=False)

    raw_values = [
        source_file,
        marketplace,
        import_timestamp,
        campaign_id,
        campaign_name,
        sku_id,
        product_title,
        date,
        to_int_like(impressions),
        to_int_like(clicks),
        f"{ad_spend:.2f}" if ad_spend else "0",
        f"{ad_sales:.2f}" if ad_sales else "0",
        to_int_like(orders),
        raw_json,
    ]

    master_values = [
        campaign_id,
        campaign_name,
        marketplace,
        sku_id,
        product_title,
        date,
        to_int_like(impressions),
        to_int_like(clicks),
        f"{ad_spend:.2f}" if ad_spend else "0",
        f"{ad_sales:.2f}" if ad_sales else "0",
        to_int_like(orders),
        ctr,
        cpc,
        roas,
        acos,
        source_file,
        import_timestamp,
    ]

    key = {
        "campaign_id": campaign_id,
        "marketplace": marketplace,
        "sku_id": sku_id,
        "date": date,
    }

    return raw_values, master_values, key


def get_existing_keys(sheets_service, spreadsheet_id: str, tab_name: str) -> Dict[Tuple[str, str, str, str], int]:
    rows = get_sheet_values(sheets_service, spreadsheet_id, f"{tab_name}!A1:Q")
    if len(rows) < 2:
        return {}

    headers = rows[0]
    campaign_index = find_header_index(headers, "Campaign_ID")
    marketplace_index = find_header_index(headers, "Marketplace")
    sku_index = find_header_index(headers, "SKU_ID")
    date_index = find_header_index(headers, "Date")
    if None in {campaign_index, marketplace_index, sku_index, date_index}:
        return {}

    existing: Dict[Tuple[str, str, str, str], int] = {}
    for row_number, row in enumerate(rows[1:], start=2):
        campaign_id = row[campaign_index].strip() if campaign_index < len(row) else ""
        marketplace = row[marketplace_index].strip() if marketplace_index < len(row) else ""
        sku_id = row[sku_index].strip() if sku_index < len(row) else ""
        date = row[date_index].strip() if date_index < len(row) else ""
        key = (campaign_id, marketplace, sku_id, date)
        if all(key) and key not in existing:
            existing[key] = row_number
    return existing


def append_rows(
    sheets_service,
    spreadsheet_id: str,
    tab_name: str,
    rows: List[List[str]],
    range_end: str,
) -> None:
    if not rows:
        return
    sheets_service.spreadsheets().values().append(
        spreadsheetId=spreadsheet_id,
        range=f"{tab_name}!A:{range_end}",
        valueInputOption="RAW",
        insertDataOption="INSERT_ROWS",
        body={"values": rows},
    ).execute()


def append_log_rows(log_rows: List[Dict[str, object]]) -> None:
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    file_exists = LOG_PATH.exists()
    with LOG_PATH.open("a", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "timestamp",
                "source_file",
                "campaign_id",
                "marketplace",
                "sku_id",
                "date",
                "status",
                "message",
            ],
        )
        if not file_exists:
            writer.writeheader()
        writer.writerows(log_rows)


def build_import_rows() -> Tuple[List[Dict[str, object]], List[Dict[str, object]]]:
    mapping = load_mapping()
    input_dir = ADS_INPUT_DIR
    input_dir.mkdir(parents=True, exist_ok=True)

    records: List[Dict[str, object]] = []
    file_reports: List[Dict[str, object]] = []

    for path in sorted(input_dir.iterdir()):
        if path.suffix.lower() not in {".csv", ".xlsx", ".xlsm"}:
            continue
        if path.name.startswith("~$"):
            continue

        headers, data_rows = read_input_file(path)
        file_reports.append({"file_name": path.name, "rows_read": len(data_rows), "rows_queued": 0})
        print(f"Found file: {path.name} | rows read: {len(data_rows)}")

        if not headers:
            print(f"  Skipped: no rows found in {path.name}")
            continue

        for row in data_rows:
            raw_values, master_values, key = normalize_row(path.name, headers, row, mapping)
            records.append({"source_file": path.name, "raw_values": raw_values, "master_values": master_values, "key": key})
            file_reports[-1]["rows_queued"] += 1

    return records, file_reports


def apply_formatting(sheets_service, spreadsheet_id: str, sheet_id: int, column_count: int) -> None:
    requests = [
        {
            "updateSheetProperties": {
                "properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 1}},
                "fields": "gridProperties.frozenRowCount",
            }
        },
        {
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1},
                "cell": {"userEnteredFormat": {"textFormat": {"bold": True}}},
                "fields": "userEnteredFormat.textFormat.bold",
            }
        },
    ]

    sheets_service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": requests}).execute()
    sheets_service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "requests": [
                {
                    "autoResizeDimensions": {
                        "dimensions": {
                            "sheetId": sheet_id,
                            "dimension": "COLUMNS",
                            "startIndex": 0,
                            "endIndex": column_count,
                        }
                    }
                }
            ]
        },
    ).execute()


def ensure_ads_import_system() -> Dict[str, object]:
    meta = load_json(SPREADSHEET_META_PATH)
    spreadsheet_id = meta["spreadsheet_id"]
    sheets_service, _, _ = build_services()

    ADS_INPUT_DIR.mkdir(parents=True, exist_ok=True)

    raw_sheet_id = ensure_tab(sheets_service, spreadsheet_id, RAW_TAB)
    master_sheet_id = ensure_tab(sheets_service, spreadsheet_id, MASTER_TAB)
    raw_headers_written = ensure_headers(sheets_service, spreadsheet_id, RAW_TAB, RAW_HEADERS)
    master_headers_written = ensure_headers(sheets_service, spreadsheet_id, MASTER_TAB, MASTER_HEADERS)

    existing_raw_keys = get_existing_keys(sheets_service, spreadsheet_id, RAW_TAB)
    existing_master_keys = get_existing_keys(sheets_service, spreadsheet_id, MASTER_TAB)

    records, file_reports = build_import_rows()

    filtered_raw_rows: List[List[str]] = []
    filtered_master_rows: List[List[str]] = []
    log_rows: List[Dict[str, object]] = []
    source_file_stats: Dict[str, Dict[str, int]] = {
        report["file_name"]: {"written_raw": 0, "written_master": 0, "skipped_duplicates": 0}
        for report in file_reports
    }

    for record in records:
        raw_values = record["raw_values"]
        master_values = record["master_values"]
        key = (master_values[0], master_values[2], master_values[3], master_values[5])
        source_file = record["source_file"]

        if key in existing_raw_keys or key in existing_master_keys:
            source_file_stats[source_file]["skipped_duplicates"] += 1
            log_rows.append(
                {
                    "timestamp": datetime.now().isoformat(timespec="seconds"),
                    "source_file": source_file,
                    "campaign_id": key[0],
                    "marketplace": key[1],
                    "sku_id": key[2],
                    "date": key[3],
                    "status": "skipped",
                    "message": "Duplicate ads key",
                }
            )
            continue

        filtered_raw_rows.append(raw_values)
        filtered_master_rows.append(master_values)
        existing_raw_keys[key] = len(existing_raw_keys) + 2
        existing_master_keys[key] = len(existing_master_keys) + 2
        source_file_stats[source_file]["written_raw"] += 1
        source_file_stats[source_file]["written_master"] += 1
        log_rows.append(
            {
                "timestamp": datetime.now().isoformat(timespec="seconds"),
                "source_file": source_file,
                "campaign_id": key[0],
                "marketplace": key[1],
                "sku_id": key[2],
                "date": key[3],
                "status": "created",
                "message": "Imported ads row",
            }
        )

    append_rows(sheets_service, spreadsheet_id, RAW_TAB, filtered_raw_rows, "N")
    append_rows(sheets_service, spreadsheet_id, MASTER_TAB, filtered_master_rows, "Q")
    append_log_rows(log_rows)

    apply_formatting(sheets_service, spreadsheet_id, raw_sheet_id, len(RAW_HEADERS))
    apply_formatting(sheets_service, spreadsheet_id, master_sheet_id, len(MASTER_HEADERS))

    print("Import summary:")
    print(f"  Files found: {len(file_reports)}")
    for report in file_reports:
        stats = source_file_stats.get(report["file_name"], {"written_raw": 0, "written_master": 0, "skipped_duplicates": 0})
        print(
            "  - "
            f"{report['file_name']}: read {report['rows_read']} rows, "
            f"wrote raw {stats['written_raw']}, wrote master {stats['written_master']}, "
            f"skipped duplicates {stats['skipped_duplicates']}"
        )
    print(f"  Rows written to ADS_RAW: {len(filtered_raw_rows)}")
    print(f"  Rows written to ADS_MASTER: {len(filtered_master_rows)}")
    print(f"  Skipped duplicates: {len([row for row in log_rows if row['status'] == 'skipped'])}")

    return {
        "spreadsheet_id": spreadsheet_id,
        "raw_tab": RAW_TAB,
        "master_tab": MASTER_TAB,
        "raw_headers_written": raw_headers_written,
        "master_headers_written": master_headers_written,
        "import_directory": str(ADS_INPUT_DIR),
        "raw_rows_imported": len(filtered_raw_rows),
        "master_rows_imported": len(filtered_master_rows),
        "files_found": len(file_reports),
        "log_path": str(LOG_PATH),
    }


def main() -> None:
    result = ensure_ads_import_system()
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
